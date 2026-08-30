"""
Load test against the real worker, on the real box.

Everything this pipeline has ever processed in production has been a fixture of
about ten rows. That tells you the code is correct and nothing at all about
whether the box survives a month-end, which is the question that matters before
five accounting firms upload at once.

So this generates workbooks at a realistic size and shape -- thousands of rows,
the same mess the fixtures have in miniature: three date conventions in one
column, numbers stored as text, a subtotal row, blank leader rows and trailing
notes -- and drives them through the queue the way an accountant would.

Run it inside the worker container, which already holds the credentials and the
dependencies:

    docker exec hermes-hermes-1 python /tmp/load_test.py --jobs 3 --rows 5000

What it measures, per job: how long the row waited to be claimed, how long each
stage took, whether anything was retried, and what the chain produced. What it
cannot measure from inside is CPU and memory, because a container sees the
host's cores rather than its own quota -- run `docker stats` on the host
alongside it and read them there.

Two deliberate choices. Jobs are enqueued through `enqueue_agent_job_internal`,
the service-role path, because the human rate limit is not what is under test
here and throttling the load generator would only measure the throttle. And
every row it writes is tagged, so `--cleanup` can find them again; nothing here
touches a dataset it did not create.
"""

from __future__ import annotations

import argparse
import io
import os
import random
import sys
import time
import uuid
from datetime import date, timedelta

import httpx
from openpyxl import Workbook

TAG = "load-test"

VENDORS = [
    "O2 Mobile", "O2 MOBILE LTD", "British Gas", "BRITISH GAS PLC", "Screwfix",
    "Screwfix Direct", "Amazon UK", "AMAZON.CO.UK", "Northwind Supplies",
    "Northwind Supplies Ltd", "Contoso Ltd", "CONTOSO LIMITED", "Travis Perkins",
    "City Electrical", "Wolseley", "Toolstation", "Rexel UK", "Edmundson",
]


def build_workbook(rows: int, seed: int) -> bytes:
    """
    A messy ledger of `rows` lines.

    Mess is the point -- a clean sheet parses in a fraction of the time and
    would report an envelope the real thing never reaches. The specific
    nastiness mirrors `make_messy_fixture.py`: mixed date conventions, numbers
    as text with currency symbols, a subtotal, and footnotes below the table.
    """
    rnd = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    # Leader rows: a title, a blank, and a stray note before the real header.
    ws["A1"] = "ACME Trading Ltd"
    ws["A2"] = f"Sales ledger export {date(2026, 8, 1):%B %Y}"
    ws["A4"] = "(figures ex VAT unless marked)"

    header = 6
    for col, name in enumerate(["Date", "Supplier", "Reference", "Net", "VAT", "Gross"], start=1):
        ws.cell(row=header, column=col, value=name)

    start = date(2026, 8, 1)
    total = 0.0
    for i in range(rows):
        r = header + 1 + i
        when = start + timedelta(days=rnd.randint(0, 27))

        # Three conventions in one column, which is what breaks naive parsers.
        style = i % 3
        if style == 0:
            ws.cell(row=r, column=1, value=when)
        elif style == 1:
            ws.cell(row=r, column=1, value=when.strftime("%d/%m/%Y"))
        else:
            ws.cell(row=r, column=1, value=when.strftime("%d-%b-%y"))

        net = round(rnd.uniform(5, 4000), 2)
        vat = round(net * 0.2, 2)
        total += net

        ws.cell(row=r, column=2, value=rnd.choice(VENDORS))
        ws.cell(row=r, column=3, value=f"INV-{rnd.randint(10000, 99999)}")

        # Every fourth amount as text with a symbol and a thousands separator.
        if i % 4 == 0:
            ws.cell(row=r, column=4, value=f"£{net:,.2f}")
        else:
            ws.cell(row=r, column=4, value=net)
        ws.cell(row=r, column=5, value=vat)
        ws.cell(row=r, column=6, value=round(net + vat, 2))

    last = header + rows
    ws.cell(row=last + 1, column=2, value="TOTAL")
    ws.cell(row=last + 1, column=4, value=round(total, 2))
    ws.cell(row=last + 3, column=1, value="Note: August includes two credit notes.")
    ws.cell(row=last + 4, column=1, value="Prepared by finance, subject to review.")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class Api:
    def __init__(self, url: str, key: str):
        self.url = url.rstrip("/")
        self.key = key
        self.http = httpx.Client(timeout=120.0)

    def _h(self, extra=None):
        h = {"apikey": self.key, "Authorization": f"Bearer {self.key}",
             "Content-Type": "application/json"}
        if extra:
            h.update(extra)
        return h

    def rpc(self, fn, params):
        r = self.http.post(f"{self.url}/rest/v1/rpc/{fn}", headers=self._h(), json=params)
        r.raise_for_status()
        return r.json()

    def select(self, table, params):
        r = self.http.get(f"{self.url}/rest/v1/{table}", headers=self._h(), params=params)
        r.raise_for_status()
        return r.json()

    def insert(self, table, row):
        r = self.http.post(f"{self.url}/rest/v1/{table}", headers=self._h({"Prefer": "return=representation"}), json=row)
        r.raise_for_status()
        return r.json()[0]

    def upload(self, bucket, path, data):
        r = self.http.post(
            f"{self.url}/storage/v1/object/{bucket}/{path}",
            headers={"apikey": self.key, "Authorization": f"Bearer {self.key}",
                     "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     "x-upsert": "true"},
            content=data,
        )
        r.raise_for_status()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--jobs", type=int, default=3, help="concurrent workbooks to submit")
    ap.add_argument("--rows", type=int, default=5000, help="rows per workbook")
    ap.add_argument("--workspace", required=True)
    ap.add_argument("--timeout", type=int, default=900)
    ap.add_argument("--cleanup", action="store_true", help="remove rows this tool created and exit")
    args = ap.parse_args()

    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SECRET_KEY", "").strip()
    if not url or not key:
        print("SUPABASE_URL and SUPABASE_SECRET_KEY must be set", file=sys.stderr)
        return 2

    api = Api(url, key)

    workspace = api.select("workspaces", {"id": f"eq.{args.workspace}", "select": "id,org_id,name"})
    if not workspace:
        print(f"workspace {args.workspace} not found", file=sys.stderr)
        return 2
    org_id = workspace[0]["org_id"]
    members = api.select("organization_members",
                         {"org_id": f"eq.{org_id}", "select": "user_id", "limit": "1"})
    if not members:
        print(f"no member found for org {org_id}", file=sys.stderr)
        return 2
    uploader = members[0]["user_id"]
    print(f"workspace {workspace[0]['name']} ({args.workspace}) uploader={uploader[:8]}")

    if args.cleanup:
        jobs = api.select("agent_jobs", {"select": "id", "payload->>note": f"eq.{TAG}"})
        print(f"{len(jobs)} tagged job(s) found; delete them from the dashboard or SQL")
        return 0

    # The worker refuses an upload that is not attached to a dataset -- a
    # dataset is the recurring thing an upload is another month of. Create one
    # for the run rather than borrowing a real one.
    dataset = api.insert("datasets", {
        "workspace_id": args.workspace,
        "name": f"load-test {int(time.time())}",
        "created_by": uploader,
    })
    print(f"dataset {dataset['id'][:8]} created for this run")

    print(f"generating {args.jobs} workbook(s) of {args.rows} rows")
    submitted = []
    for n in range(args.jobs):
        blob = build_workbook(args.rows, seed=n)
        upload_id = str(uuid.uuid4())
        path = f"{org_id}/{args.workspace}/{upload_id}.xlsx"

        t0 = time.perf_counter()
        api.upload("raw", path, blob)
        upload_seconds = time.perf_counter() - t0

        row = api.insert("raw_uploads", {
            "id": upload_id, "workspace_id": args.workspace, "storage_path": path,
            "original_filename": f"load-{args.rows}-{n}.xlsx",
            "byte_size": len(blob), "dataset_id": dataset["id"],
            # `stored`, not `uploaded` -- the enum is (pending, stored, failed).
            "status": "stored",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            # NOT NULL: the row records who sent the file, and the load test
            # borrows the workspace's own member so lineage stays truthful.
            "uploaded_by": uploader,
        })

        job = api.rpc("enqueue_agent_job_internal", {
            "p_workspace_id": args.workspace, "p_kind": "parse_workbook",
            "p_payload": {"note": TAG}, "p_raw_upload_id": row["id"], "p_priority": 100,
        })
        submitted.append(job["id"])
        print(f"  #{n}: {len(blob)/1024:.0f} KB uploaded in {upload_seconds:.1f}s -> job {job['id'][:8]}")

    print(f"\nwaiting (timeout {args.timeout}s). Run `docker stats` on the host now.\n")
    deadline = time.time() + args.timeout
    seen: dict[str, str] = {}
    started = time.time()

    while time.time() < deadline:
        rows = api.select("agent_jobs", {
            "select": "id,kind,status,attempts,claimed_by,created_at,started_at,finished_at,error",
            "payload->>note": f"eq.{TAG}",
            "order": "created_at.asc",
        })
        for r in rows:
            state = f"{r['kind']}:{r['status']}"
            if seen.get(r["id"]) != state:
                seen[r["id"]] = state
                print(f"  [{time.time()-started:6.1f}s] {r['id'][:8]} {r['kind']:<18} "
                      f"{r['status']:<10} attempt={r['attempts']} {r.get('error') or ''}")
        if rows and all(r["status"] in ("succeeded", "failed") for r in rows):
            # The chain adds jobs as it goes, so wait a beat for stragglers.
            time.sleep(6)
            rows = api.select("agent_jobs", {
                "select": "id,kind,status,attempts,created_at,started_at,finished_at",
                "payload->>note": f"eq.{TAG}", "order": "created_at.asc"})
            if all(r["status"] in ("succeeded", "failed") for r in rows):
                break
        time.sleep(3)

    print("\n=== results ===")
    rows = api.select("agent_jobs", {
        "select": "id,kind,status,attempts,created_at,started_at,finished_at,error",
        "payload->>note": f"eq.{TAG}", "order": "created_at.asc"})

    from datetime import datetime

    def parse(t):
        return datetime.fromisoformat(t.replace("Z", "+00:00")) if t else None

    ok = fail = 0
    for r in rows:
        created, started_at, finished = parse(r["created_at"]), parse(r["started_at"]), parse(r["finished_at"])
        queued = (started_at - created).total_seconds() if started_at else None
        ran = (finished - started_at).total_seconds() if finished and started_at else None
        ok += r["status"] == "succeeded"
        fail += r["status"] == "failed"
        print(f"  {r['kind']:<18} {r['status']:<10} attempts={r['attempts']} "
              f"queued={queued if queued is None else f'{queued:6.1f}s'} "
              f"ran={ran if ran is None else f'{ran:7.1f}s'} {r.get('error') or ''}")

    print(f"\n  {ok} succeeded, {fail} failed, {len(rows)} total")
    print(f"  retries: {sum(max(0, r['attempts'] - 1) for r in rows)}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
