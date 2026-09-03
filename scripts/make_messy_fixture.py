"""
Builds the deliberately messy XLSX fixtures.

PRD section 6 lists what real accountant spreadsheets actually look like, and
calls the parser a P0 deliverable rather than an assumption. Every trait in the
August file is one of the failure modes listed there.

There are two months, and the second is the more important one. Criterion 6 of
the MVP is "a second month's file auto-matches the recipe and replays it", and
that cannot be tested with one file. September therefore keeps August's layout
exactly -- same headers, same header row, same types, so it produces the same
source_signature and matches the recipe -- while changing everything a real
second month would change:

  * different transactions, and more of them
  * a supplier spelled a new way ("Northwind Supplies Limited"), which should
    be offered as an ambiguous match rather than silently merged
  * a supplier nobody has seen before ("Litware Inc"), which should be
    reported as new rather than guessed at
  * no duplicate rows, because last month's duplicate was a one-off
  * totals that reconcile, so the blocking finding does not recur

Between them the two files are the eval harness for the whole loop.

Usage: python scripts/make_messy_fixture.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

OUT_AUGUST = Path("fixtures/messy/acme-sales-2026-08.xlsx")
OUT_SEPTEMBER = Path("fixtures/messy/acme-sales-2026-09.xlsx")

# The same August ledger in the legacy binary format, for the reader that has to
# handle it. A customer exporting from an older system sends this, and until it
# parsed, the first thing they saw was a refusal.
#
# Written with xlwt, which is a *development* dependency and not in
# requirements.txt: the agent only ever reads .xls, and adding a writer to the
# production image to build a test file would be paying for it in every
# deployment. The fixture it produces is committed, so the test suite needs
# neither xlwt nor this script.
OUT_AUGUST_XLS = Path("fixtures/messy/acme-sales-2026-08.xls")


def build() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Aug"

    # Title block above the data: the header row is not row 1.
    ws["A1"] = "ACME Trading Ltd"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")  # merged cells
    ws["A2"] = "Sales export - August 2026"
    ws.merge_cells("A2:E2")
    ws["A3"] = "Generated 01/09/2026 by Sage 50"
    # row 4 deliberately blank

    # The real header row, on row 5.
    headers = ["Date", "Invoice", "Supplier", "Net Sales", "VAT"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows. Dates arrive in three conventions in one column: real datetimes
    # written as Excel serials, DD/MM/YYYY text, and MM/DD/YYYY text. Numbers
    # arrive as text with thousands separators, currency symbols and
    # parentheses negatives.
    rows = [
        ["01/08/2026", "INV-1001", "Northwind Supplies Ltd", "1,240.00", "248.00"],
        ["02/08/2026", "INV-1002", "northwind supplies", "£880.50", "176.10"],
        ["08/03/2026", "INV-1003", "Contoso Ltd.", "2,015.75", "403.15"],  # MM/DD
        [None, None, None, None, None],  # blank separator row
        ["04/08/2026", "INV-1004", "CONTOSO LIMITED", "(150.00)", "(30.00)"],  # credit note
        ["05/08/2026", "INV-1005", "Fabrikam  Ltd", "3,420.10", "684.02"],
        ["Subtotal", None, None, "7,406.35", "1,481.27"],  # embedded subtotal row
        [None, None, None, None, None],
        ["06/08/2026", "INV-1006", "Fabrikam Ltd", "965.00", "193.00"],
        ["07/08/2026", "INV-1007", "Tailspin Toys", "1,200", "240"],
        ["07/08/2026", "INV-1007", "Tailspin Toys", "1,200", "240"],  # exact duplicate
        ["09/08/2026", "INV-1008", "Wide World Importers", "-410.25", "-82.05"],
    ]

    for offset, row in enumerate(rows, start=6):
        for col, value in enumerate(row, start=1):
            ws.cell(row=offset, column=col, value=value)

    # Trailing total row that must not be treated as a transaction.
    total_row = 6 + len(rows) + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="10,361.35")
    ws.cell(row=total_row, column=5, value="2,072.27")

    # Footnotes and disclaimer text below the data.
    ws.cell(row=total_row + 2, column=1, value="* Excludes intercompany transfers.")
    ws.cell(
        row=total_row + 3,
        column=1,
        value="This report is provided for information only and is not audited.",
    )

    # A second sheet, because one file routinely holds several tables.
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Vendor code changes"
    ws2["A3"] = "Old code"
    ws2["B3"] = "New code"
    ws2["A4"] = "NW-01"
    ws2["B4"] = "NORTH-001"
    ws2["A5"] = "CTS-04"
    ws2["B5"] = "CONT-004"

    return wb


def build_september() -> Workbook:
    """
    Month two: same shape, different content.

    The layout is copied deliberately rather than varied. source_signature is
    computed from column names, types and header position, so any change here
    would stop the recipe matching -- and a test that fails because the fixture
    drifted teaches nothing about the code.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Sep"

    ws["A1"] = "ACME Trading Ltd"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Sales export - September 2026"
    ws.merge_cells("A2:E2")
    ws["A3"] = "Generated 01/10/2026 by Sage 50"
    # row 4 blank, header on row 5 -- same as August.

    headers = ["Date", "Invoice", "Supplier", "Net Sales", "VAT"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ["01/09/2026", "INV-1009", "Northwind Supplies Ltd", "2,150.00", "430.00"],
        # A spelling of Northwind that has not been seen before. It folds to the
        # same entity key, so the agent should offer it as a match and ask --
        # not merge it silently.
        ["02/09/2026", "INV-1010", "Northwind Supplies Limited", "1,890.25", "378.05"],
        ["03/09/2026", "INV-1011", "Contoso Ltd.", "3,410.00", "682.00"],
        [None, None, None, None, None],
        ["04/09/2026", "INV-1012", "Fabrikam Ltd", "£975.40", "195.08"],
        # Genuinely new. There is nothing to match it against, and guessing
        # would invent a relationship that does not exist.
        ["05/09/2026", "INV-1013", "Litware Inc", "4,200.00", "840.00"],
        ["Subtotal", None, None, "12,625.65", "2,525.13"],
        [None, None, None, None, None],
        ["08/09/2026", "INV-1014", "Tailspin Toys", "1,150.00", "230.00"],
        ["09/09/2026", "INV-1015", "Wide World Importers", "(220.50)", "(44.10)"],
        ["10/09/2026", "INV-1016", "Contoso Ltd.", "2,780.00", "556.00"],
        ["11/09/2026", "INV-1017", "Fabrikam Ltd", "1,640.75", "328.15"],
    ]

    for offset, row in enumerate(rows, start=6):
        for col, value in enumerate(row, start=1):
            ws.cell(row=offset, column=col, value=value)

    # A trailing total that actually reconciles this month: 17,975.90.
    total_row = 6 + len(rows) + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="17,975.90")
    ws.cell(row=total_row, column=5, value="3,595.18")

    ws.cell(row=total_row + 2, column=1, value="* Excludes intercompany transfers.")
    ws.cell(
        row=total_row + 3,
        column=1,
        value="This report is provided for information only and is not audited.",
    )

    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Vendor code changes"
    ws2["A3"] = "Old code"
    ws2["B3"] = "New code"
    ws2["A4"] = "LW-09"
    ws2["B4"] = "LITW-009"

    return wb


def build_august_xls(path: Path) -> None:
    """
    August again, as a .xls.

    Deliberately not a byte-for-byte twin of the .xlsx. It carries the traits
    the *reader* has to get right, which are not the same traits the structure
    detector is tested on: real date cells (which xlrd hands back as floats plus
    an epoch), a boolean, a formula error cell, a hidden sheet that must not be
    parsed as the main table, and a TOTAL row that must not become a
    transaction.
    """
    import datetime as dt

    import xlwt

    book = xlwt.Workbook()
    sheet = book.add_sheet("August")

    # A title block above the header, as every real export has.
    sheet.write(0, 0, "ACME TRADING LTD")
    sheet.write(1, 0, "Sales ledger — August 2026")

    for column, name in enumerate(["Date", "Invoice", "Supplier", "Net Sales", "VAT", "Paid"]):
        sheet.write(3, column, name)

    dates = xlwt.XFStyle()
    dates.num_format_str = "DD/MM/YYYY"

    rows = [
        (dt.datetime(2026, 8, 3), "INV-1001", "Contoso Ltd", 1240.50, 248.10, True),
        (dt.datetime(2026, 8, 7), "INV-1002", "CONTOSO LIMITED", 980.00, 196.00, True),
        (dt.datetime(2026, 8, 11), "INV-1003", "Fabrikam", 2015.75, 403.15, False),
        (dt.datetime(2026, 8, 19), "INV-1004", "Northwind Supplies", 610.25, 122.05, True),
    ]
    for index, (date, invoice, supplier, net, vat, paid) in enumerate(rows, start=4):
        sheet.write(index, 0, date, dates)
        sheet.write(index, 1, invoice)
        sheet.write(index, 2, supplier)
        sheet.write(index, 3, net)
        sheet.write(index, 4, vat)
        sheet.write(index, 5, paid)

    # A formula that evaluates to an error. xlrd reports it as an error cell,
    # and the reader must return None rather than the error's numeric code --
    # which would otherwise be inferred as a number in a boolean column.
    sheet.write(8, 5, xlwt.Formula("1/0"))

    # The trailing total, which must never be read as a transaction.
    sheet.write(9, 2, "TOTAL")
    sheet.write(9, 3, 4846.50)
    sheet.write(9, 4, 969.30)

    # A hidden lookup sheet. Parsing this as the main table is a classic wrong
    # answer, so the reader skips it.
    lookup = book.add_sheet("Lookup")
    lookup.write(0, 0, "supplier")
    lookup.write(1, 0, "do not parse me")
    lookup.visibility = 1

    book.save(str(path))


if __name__ == "__main__":
    OUT_AUGUST.parent.mkdir(parents=True, exist_ok=True)

    for path, workbook in ((OUT_AUGUST, build()), (OUT_SEPTEMBER, build_september())):
        workbook.save(path)
        print(f"wrote {path} ({path.stat().st_size} bytes)")

    try:
        build_august_xls(OUT_AUGUST_XLS)
        print(f"wrote {OUT_AUGUST_XLS} ({OUT_AUGUST_XLS.stat().st_size} bytes)")
    except ImportError:
        print("skipped the .xls fixture: pip install xlwt to rebuild it")
