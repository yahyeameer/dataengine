"""
Upload to download, end to end, including opening the file at the other end.

The brief for this feature is blunt about one thing: do not report success
because the model returned results. The success state is a workbook that exists,
opens, holds the original data and carries the categories -- and the only way to
test that claim is to take the bytes the job uploaded and open them.

So the centre of this file is `test_the_downloaded_workbook_holds_the_categories`.
Everything above it exercises the pieces that decide what goes into that file;
everything below it exercises the ways it is allowed to fail.

The fakes are the smallest thing that lets the job run: an in-memory Parquet
file, and a Supabase that keeps what the job writes instead of storing it. The
storage fake matters more than it looks -- it is what lets the test reach the
exact bytes that would have been served to the browser.
"""

from __future__ import annotations

import io
from typing import Any

import polars as pl
import pytest
from openpyxl import load_workbook

from hermes.config import Config, GovUKConfig, LLMConfig
from hermes.jobs import JobContext, JobError, handle_categorise_statement
from hermes.tools import autopilot, hmrc


# A month of a real-looking UK current account. Deliberately mixed: merchants the
# rules know cold, a marketplace line that cannot be known, private spending, a
# transfer, a drawing, and one trader nothing recognises.
_STATEMENT: list[tuple[str, str, float]] = [
    ("2026-04-02", "CARD PAYMENT TO TFL TRAVEL CH 02APR CARD 1234", 18.40),
    ("2026-04-03", "DIRECT DEBIT BRITISH GAS REF 8823991", 96.12),
    ("2026-04-05", "O2 UK LTD", 32.00),
    ("2026-04-06", "AMZN MKTP UK*2H4XY9", 54.99),
    ("2026-04-07", "SCREWFIX DIRECT 0500414141", 212.80),
    ("2026-04-09", "HMRC VAT 1234567890", 1840.00),
    ("2026-04-10", "TRANSFER TO SAVINGS", 500.00),
    ("2026-04-11", "TESCO STORES 3411", 63.25),
    ("2026-04-12", "XERO SUBSCRIPTION", 33.00),
    ("2026-04-14", "CASH WITHDRAWAL LINK ATM", 200.00),
    ("2026-04-15", "PARKING FINE PCN 88213", 65.00),
    ("2026-04-16", "QRZ TRADING LTD 8891", 410.00),
]

# Every distinct description except the one nothing recognises.
_BY_RULE = len(_STATEMENT) - 1
_UNPLACEABLE = "QRZ TRADING LTD 8891"


def _parquet() -> bytes:
    frame = pl.DataFrame(
        {
            "__source_row": list(range(2, len(_STATEMENT) + 2)),
            "Date": [row[0] for row in _STATEMENT],
            "Description": [row[1] for row in _STATEMENT],
            "Amount": [row[2] for row in _STATEMENT],
        }
    )
    buffer = io.BytesIO()
    frame.write_parquet(buffer)
    return buffer.getvalue()


class _Stored:
    def __init__(self, bucket: str, path: str, data: bytes):
        self.bucket, self.path, self.size = bucket, path, len(data)
        self.data = data


class _OfflineLLM:
    """An agent host with no API key. The rules must carry the run alone."""

    enabled = False

    def categorize_values(self, *a: Any, **k: Any):  # pragma: no cover - must not run
        raise AssertionError("the rules path asked a model that is not configured")


class _FakeSupabase:
    """Records the RPCs and keeps the objects, so a test can open the export."""

    def __init__(self, parquet: bytes | None = None):
        self.rpc_calls: list[tuple[str, dict[str, Any]]] = []
        self.objects: dict[str, _Stored] = {}
        self.parquet = parquet if parquet is not None else _parquet()

    # -- reads ----------------------------------------------------------------

    def select(self, table: str, **kwargs: Any) -> list[dict[str, Any]]:
        if table == "dataset_versions":
            return [
                {
                    "id": "version-1",
                    "dataset_id": "dataset-1",
                    "version_no": 1,
                    "kind": "cleaned",
                    "parquet_path": "org/ws/2026-04/dataset-1__parse.parquet",
                    "row_count": len(_STATEMENT),
                    "parent_version_id": None,
                    "raw_upload_id": "upload-1",
                    "produced_by_run_id": None,
                }
            ]
        if table == "datasets":
            return [{"id": "dataset-1", "name": "April statement"}]
        if table == "raw_uploads":
            return [{"id": "upload-1", "original_filename": "april-statement.xlsx"}]
        return []

    def download(self, bucket: str, path: str, max_bytes: int) -> bytes:
        return self.parquet

    # -- writes ---------------------------------------------------------------

    def upload(self, bucket: str, path: str, data: bytes, **kwargs: Any) -> _Stored:
        stored = _Stored(bucket, path, data)
        self.objects[path] = stored
        return stored

    def rpc(self, function: str, params: dict[str, Any] | None = None) -> Any:
        self.rpc_calls.append((function, params or {}))
        if function == "record_dataset_version":
            return {"id": "version-2", "version_no": 2}
        if function == "append_proposed_changes":
            return len((params or {}).get("p_proposals", []))
        if function == "record_hmrc_source":
            return {"source_id": "source-1", "changed": False}
        return None

    # -- assertions -----------------------------------------------------------

    def called(self, name: str) -> dict[str, Any] | None:
        for function, params in self.rpc_calls:
            if function == name:
                return params
        return None

    def proposal(self) -> dict[str, Any]:
        params = self.called("append_proposed_changes")
        assert params is not None, "no proposal was written"
        return params["p_proposals"][0]


def _context(
    supabase: Any,
    payload: dict[str, Any] | None = None,
    llm: Any | None = None,
    govuk: GovUKConfig | None = None,
) -> JobContext:
    return JobContext(
        config=Config(
            supabase_url="https://example.supabase.co",
            service_key="test",
            worker_id="test",
            hostname="test",
            llm=LLMConfig(),
            govuk=govuk or GovUKConfig(),
        ),
        supabase=supabase,
        llm=llm or _OfflineLLM(),
        job={
            "id": "job-1",
            "workspace_id": "ws-1",
            "org_id": "org-1",
            "dataset_version_id": "version-1",
            "payload": payload or {},
            "requested_by": None,
        },
        heartbeat=lambda progress: None,
    )


def _run(**kwargs: Any) -> tuple[_FakeSupabase, dict[str, Any]]:
    supabase = kwargs.pop("supabase", None) or _FakeSupabase()
    return supabase, handle_categorise_statement(_context(supabase, **kwargs))


def _export_sheet(supabase: _FakeSupabase, result: dict[str, Any]) -> list[list[Any]]:
    """The workbook the customer would have downloaded, as a grid."""
    stored = supabase.objects[result["export_path"]]
    workbook = load_workbook(io.BytesIO(stored.data), read_only=True, data_only=True)
    try:
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


# -----------------------------------------------------------------------------
# The file at the other end
# -----------------------------------------------------------------------------


def test_the_downloaded_workbook_holds_the_categories():
    """
    The acceptance criterion, tested the way the brief asks for it.

    Not "the job returned a path" and not "the mapping was correct" -- the actual
    bytes that were uploaded, opened with the library Excel users effectively
    are, and read row by row.
    """
    supabase, result = _run()
    grid = _export_sheet(supabase, result)

    header = grid[0]
    body = grid[1:]

    # The original data is still there, unchanged and in its original columns.
    assert header[:3] == ["Date", "Description", "Amount"]
    assert len(body) == len(_STATEMENT)
    assert [row[1] for row in body] == [line[1] for line in _STATEMENT]
    assert [row[2] for row in body] == [line[2] for line in _STATEMENT]

    # And the three new columns are beside it, not instead of it.
    for column in hmrc.OUTPUT_COLUMNS:
        assert column in header

    category_at = header.index(hmrc.CATEGORY_COLUMN)
    box_at = header.index(hmrc.BOX_COLUMN)

    by_description = {row[1]: row for row in body}

    assert by_description["CARD PAYMENT TO TFL TRAVEL CH 02APR CARD 1234"][category_at] == (
        "Travel and Subsistence"
    )
    assert by_description["CARD PAYMENT TO TFL TRAVEL CH 02APR CARD 1234"][box_at] == (
        "SA103F box 20"
    )
    assert by_description["DIRECT DEBIT BRITISH GAS REF 8823991"][category_at] == (
        "Premises Costs"
    )
    assert by_description["HMRC VAT 1234567890"][category_at] == "HMRC and Tax Payments"
    assert by_description["TRANSFER TO SAVINGS"][category_at] == "Transfers Between Accounts"
    assert by_description["CASH WITHDRAWAL LINK ATM"][category_at] == "Owner Drawings"

    # Every row has a category. A blank would be the failure that looks like
    # success right up until somebody filters the column.
    assert all(row[category_at] for row in body)


def test_the_uncertain_rows_say_so_rather_than_claiming_a_deduction():
    supabase, result = _run()
    grid = _export_sheet(supabase, result)
    header, body = grid[0], grid[1:]

    category_at = header.index(hmrc.CATEGORY_COLUMN)
    confidence_at = header.index(hmrc.CONFIDENCE_COLUMN)
    by_description = {row[1]: row for row in body}

    # A marketplace line names a shop, not a purchase.
    assert by_description["AMZN MKTP UK*2H4XY9"][category_at] == hmrc.FALLBACK
    assert by_description["AMZN MKTP UK*2H4XY9"][confidence_at] == "Needs review"

    # A supermarket is private until something says otherwise.
    assert by_description["TESCO STORES 3411"][category_at] == "Personal, Non-Business"

    # A fine is never allowable, however much it looks like a parking charge --
    # and note it is *not* flagged. "Flagged" means the agent is unsure, and it
    # is not unsure here: it is certain this is not deductible. Confidence is
    # about the strength of the evidence, not about whether the answer is
    # welcome.
    assert by_description["PARKING FINE PCN 88213"][category_at] == "Personal, Non-Business"
    assert by_description["PARKING FINE PCN 88213"][confidence_at] == "High"

    # And the trader nothing recognises is flagged, not guessed at.
    assert by_description[_UNPLACEABLE][category_at] == hmrc.FALLBACK

    # The marketplace, the supermarket and the unknown trader.
    assert result["rows_flagged"] == 3
    assert result["rows_categorised"] == len(_STATEMENT) - 3
    assert result["rows_total"] == len(_STATEMENT)


def test_a_box_never_disagrees_with_its_category():
    # The reason the three columns are written by one operation rather than
    # three. A box that belongs to a different category is a filing error that
    # would survive every other check in this file.
    supabase, result = _run()
    grid = _export_sheet(supabase, result)
    header, body = grid[0], grid[1:]

    category_at = header.index(hmrc.CATEGORY_COLUMN)
    box_at = header.index(hmrc.BOX_COLUMN)

    for row in body:
        assert (row[box_at] or "") == hmrc.BOX_BY_CATEGORY.get(row[category_at], "")


def test_it_runs_with_no_model_configured():
    _supabase, result = _run()
    assert result["model_used"] is None
    assert result["values_by_rule"] == _BY_RULE
    assert result["validated"] is True


# -----------------------------------------------------------------------------
# The record behind the file
# -----------------------------------------------------------------------------


def test_the_change_is_proposed_approved_and_applied_in_that_order():
    """
    Hiding the workflow from the customer must not remove it from the record.

    The customer never sees a review queue, and every step it used to make them
    click still happens: a proposal with its evidence, an approval, a new
    version whose parent is the old one, and the proposal marked applied.
    """
    supabase, result = _run()
    order = [name for name, _ in supabase.rpc_calls]

    assert order.index("append_proposed_changes") < order.index(
        "auto_approve_proposed_changes"
    )
    assert order.index("auto_approve_proposed_changes") < order.index(
        "record_dataset_version"
    )
    assert "mark_changes_applied" in order

    version = supabase.called("record_dataset_version")
    assert version["p_parent_version_id"] == "version-1"
    assert version["p_metadata"]["auto_applied"] is True
    assert result["parent_version_id"] == "version-1"


def test_the_automatic_approval_says_it_was_automatic():
    # The whole reason this does not simply update the row: an approval nobody
    # made must not be indistinguishable in the audit log from one somebody did.
    supabase, _ = _run()
    approval = supabase.called("auto_approve_proposed_changes")

    assert approval is not None
    assert "automatically" in approval["p_note"].lower()
    assert approval["p_group_keys"] == ["hmrc:Description"]


def test_every_classification_carries_its_reason():
    supabase, _ = _run()
    evidence = supabase.proposal()["evidence"]

    assert evidence["taxonomy"] == hmrc.TAXONOMY
    assert evidence["boxes"]["Travel and Subsistence"] == "SA103F box 20"

    # Not a sample. Every value, with why it went where it went, so a reviewer
    # who wants to check the four hundred they were not shown can.
    reasoning = evidence["reasoning"]
    assert len(reasoning) == len(_STATEMENT)

    tfl = reasoning["card payment to tfl travel ch 02apr card 1234"]
    assert tfl["category"] == "Travel and Subsistence"
    assert tfl["confidence"] == hmrc.HIGH
    assert tfl["source"] == "rule"
    assert "Transport for London" in tfl["evidence"] or "transport" in tfl["evidence"].lower()

    amazon = reasoning["amzn mktp uk*2h4xy9"]
    assert amazon["confidence"] == hmrc.LOW
    assert "does not say" in amazon["evidence"]


def test_the_proposal_stays_at_the_review_tier():
    # Applying automatically does not promote the finding. The tier is what any
    # other path in the product reads to decide what it may do unattended, and a
    # judgement about somebody else books is not an Auto-tier fact.
    supabase, _ = _run()
    assert supabase.proposal()["confidence"] == "medium"


def test_the_export_is_named_after_the_file_the_client_sent():
    _supabase, result = _run()
    assert result["source_filename"] == "april-statement.xlsx"
    assert result["export_path"].endswith("__v2__categorised.xlsx")
    assert result["bucket"] == "exports"


# -----------------------------------------------------------------------------
# The model, when there is one
# -----------------------------------------------------------------------------


def test_the_model_sees_only_what_the_rules_could_not_place():
    seen: dict[str, Any] = {}

    class _Tail:
        enabled = True

        def categorize_values(self, column, values, categories=None, hint=None):
            seen["values"] = list(values)
            seen["categories"] = list(categories or [])
            return (
                {_UNPLACEABLE.lower(): "Cost of Goods Bought for Resale"},
                [],
                "test-model",
                None,
            )

    supabase, result = _run(llm=_Tail())

    assert seen["values"] == [_UNPLACEABLE]
    assert seen["categories"] == list(hmrc.CATEGORY_NAMES)
    assert result["values_by_model"] == 1

    grid = _export_sheet(supabase, result)
    header, body = grid[0], grid[1:]
    row = next(row for row in body if row[1] == _UNPLACEABLE)
    assert row[header.index(hmrc.CATEGORY_COLUMN)] == "Cost of Goods Bought for Resale"
    # Medium, not high -- the model answered because the evidence was thin.
    assert row[header.index(hmrc.CONFIDENCE_COLUMN)] == "Medium"


def test_a_rule_beats_the_model_where_they_disagree():
    class _Contrarian:
        enabled = True

        def categorize_values(self, column, values, categories=None, hint=None):
            # Answering about a value it was never shown. The router drops those
            # itself; this is the second line of defence, and it matters because
            # a rule is reviewable prose and a model answer is not.
            return {"hmrc vat 1234567890": "Office Costs"}, [], "test-model", None

    supabase, result = _run(llm=_Contrarian())
    grid = _export_sheet(supabase, result)
    header, body = grid[0], grid[1:]

    row = next(row for row in body if row[1] == "HMRC VAT 1234567890")
    assert row[header.index(hmrc.CATEGORY_COLUMN)] == "HMRC and Tax Payments"


def test_a_model_failure_does_not_lose_what_the_rules_decided():
    class _Broken:
        enabled = True

        def categorize_values(self, *a: Any, **k: Any):
            return {}, [], "test-model", "429 rate limited"

    supabase, result = _run(llm=_Broken())

    assert result["values_by_rule"] == _BY_RULE
    assert result["rows_total"] == len(_STATEMENT)
    # The one value nobody placed is flagged rather than lost.
    grid = _export_sheet(supabase, result)
    header, body = grid[0], grid[1:]
    row = next(row for row in body if row[1] == _UNPLACEABLE)
    assert row[header.index(hmrc.CATEGORY_COLUMN)] == hmrc.FALLBACK


# -----------------------------------------------------------------------------
# Refusing rather than producing something wrong
# -----------------------------------------------------------------------------


def test_a_file_with_no_descriptions_is_refused_in_plain_words():
    frame = pl.DataFrame(
        {
            "__source_row": [2, 3, 4],
            "Opening": [1.0, 2.0, 3.0],
            "Closing": [2.0, 3.0, 4.0],
        }
    )
    buffer = io.BytesIO()
    frame.write_parquet(buffer)

    supabase = _FakeSupabase(parquet=buffer.getvalue())

    with pytest.raises(JobError) as caught:
        handle_categorise_statement(_context(supabase))

    message = str(caught.value)
    assert "could not find" in message.lower()
    # Nothing technical reaches the accountant.
    for forbidden in ("parquet", "column_hash", "supabase", "traceback"):
        assert forbidden not in message.lower()


def test_a_workbook_that_fails_its_own_checks_is_never_offered():
    """
    The gate that makes "your file is ready" mean something.

    Simulated by corrupting the workbook between writing and validating, which
    is the shape of every real version of this failure -- a serialiser bug, a
    truncated upload, a column that silently did not get written.
    """
    supabase = _FakeSupabase()
    context = _context(supabase)

    original = autopilot.validate_export

    def _fail(*a: Any, **k: Any):
        raise autopilot.ValidationError("row 4 lost its original 'Description' value")

    autopilot.validate_export = _fail
    try:
        with pytest.raises(JobError) as caught:
            handle_categorise_statement(context)
    finally:
        autopilot.validate_export = original

    assert "did not pass our own checks" in str(caught.value)
    assert caught.value.retryable is False
    # No export object was written for anybody to find.
    assert not [path for path in supabase.objects if "categorised" in path]


def test_an_empty_file_is_refused():
    frame = pl.DataFrame({"__source_row": [], "Description": []})
    buffer = io.BytesIO()
    frame.write_parquet(buffer)

    with pytest.raises(JobError) as caught:
        handle_categorise_statement(_context(_FakeSupabase(parquet=buffer.getvalue())))

    assert "no transactions" in str(caught.value)


# -----------------------------------------------------------------------------
# Choosing the column
# -----------------------------------------------------------------------------


def test_it_finds_the_description_among_plausible_neighbours():
    rows = [
        {"Date": "2026-04-02", "Type": "DD", "Transaction": line[1], "Out": line[2]}
        for line in _STATEMENT
    ]
    assert autopilot.choose_description_column(list(rows[0].keys()), rows) == "Transaction"


def test_it_finds_the_description_even_with_no_usable_column_names():
    # What a PDF-to-Excel conversion produces. The names say nothing, so the only
    # honest signal is whether the HMRC rules recognise the values.
    rows = [
        {"Col1": line[0], "Col2": line[1], "Col3": line[2]} for line in _STATEMENT
    ]
    assert autopilot.choose_description_column(list(rows[0].keys()), rows) == "Col2"


def test_it_refuses_rather_than_categorising_the_date_column():
    rows = [{"Date": line[0], "Amount": line[2]} for line in _STATEMENT]
    assert autopilot.choose_description_column(list(rows[0].keys()), rows) is None


def test_a_named_column_overrides_the_guess():
    supabase, result = _run(payload={"column": "Description"})
    assert result["source_column"] == "Description"
