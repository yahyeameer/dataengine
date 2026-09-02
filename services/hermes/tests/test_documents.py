"""
Client-facing report format tests.

These cover the failures that a byte count cannot see. A PDF that renders a
black box where an arrow should be is still a valid PDF; a workbook whose
totals are text is still a valid workbook, and the client discovers it by
selecting a column and getting nothing in the status bar. So the assertions
here are about the *content* of the file rather than its existence: that
numbers survive as numbers, that a credit is still negative on the other side,
that a name with an ampersand in it does not take the build down, and that a
brand colour nobody validated cannot fail a month-end job.

No database and no network, like the other tool tests.
"""

from __future__ import annotations

import datetime as dt
import io
import zipfile

from openpyxl import load_workbook

from hermes.tools import documents
from hermes.tools.report import Table, build_report_document

KPIS = {
    "row_count": 12483,
    "period": {"earliest": "2026-01-01", "latest": "2026-01-31"},
    "amount": {"total": 184320.55, "negative_rows": 8},
    "monthly": {
        "metric": "amount",
        "series": [
            {"month": "2025-12", "total": 163900.5},
            {"month": "2026-01", "total": 184320.55},
        ],
    },
    "top_by_route": [
        {"label": "Mombasa - Nairobi", "total": 84200.0, "rows": 410},
        {"label": "Smith & Sons <Holdings>", "total": -12400.0, "rows": 88},
    ],
}

SIGNALS = {
    "declared_totals": {
        "checked": True,
        "all_reconcile": False,
        "checks": [
            {
                "column": "amount",
                "reconciles": False,
                "computed": 184320.55,
                "declared": 184000.0,
                "difference": 320.55,
            }
        ],
    },
    "exact_duplicates": {"duplicate_rows": 12},
    "entity_variants": {"columns": []},
}


def _document():
    return build_report_document(
        workspace_name="Kentex Cargo",
        dataset_name="January shipments",
        version_no=3,
        kpis=KPIS,
        profile_signals=SIGNALS,
        comparison={
            "metric": "amount",
            "total_a": 163900.5,
            "total_b": 184320.55,
            "difference": 20420.05,
            "percent_change": 12.46,
            "drivers": [],
        },
        provenance={"source_filename": "Kentex_Jan_2026.xlsx", "parent_version_no": 2},
        narrative="Revenue rose 12.5% on December.",
        generated_at=dt.datetime(2026, 9, 2, 14, 30, tzinfo=dt.timezone.utc),
    )


def test_every_format_produces_a_file_of_its_own_type():
    document = _document()
    seen = {}
    for fmt in documents.FORMATS:
        payload, content_type, extension = documents.render_document(document, fmt)
        assert extension == fmt
        assert content_type == documents.CONTENT_TYPES[fmt]
        assert len(payload) > 500
        seen[fmt] = payload

    assert seen["pdf"].startswith(b"%PDF-")
    # docx and xlsx are zip containers; a truncated one still starts with PK.
    for fmt, member in (("docx", "word/document.xml"), ("xlsx", "xl/workbook.xml")):
        with zipfile.ZipFile(io.BytesIO(seen[fmt])) as archive:
            assert archive.testzip() is None
            assert member in archive.namelist()


def test_an_unknown_format_falls_back_to_markdown_rather_than_failing():
    payload, content_type, extension = documents.render_document(_document(), "pptx")
    assert extension == "md"
    assert content_type.startswith("text/markdown")
    assert payload.startswith(b"# January shipments")


def test_the_workbook_holds_numbers_not_formatted_text():
    payload, _type, _extension = documents.render_document(_document(), "xlsx")
    sheet = load_workbook(io.BytesIO(payload)).active

    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert 184320.55 in values, "the headline total should be a number Excel can sum"
    assert 12483 in values or 12483.0 in values
    # The credit keeps its sign rather than arriving as the string "(£12,400.00)".
    assert -12400.0 in values

    formatted = [
        cell.number_format
        for row in sheet.iter_rows()
        for cell in row
        if cell.value == 184320.55
    ]
    assert any("£" in fmt for fmt in formatted), "money should carry a currency format"


def test_the_workbook_keeps_the_client_name_and_a_readable_ink():
    brand = documents.Brand.for_client(name="Pale Gold Ltd", accent="#f2d675")
    payload, _type, _extension = documents.render_document(_document(), "xlsx", brand)
    sheet = load_workbook(io.BytesIO(payload)).active

    banner = sheet.cell(row=2, column=2)
    assert banner.value == "Pale Gold Ltd"
    # Near-black on a pale gold band, not white on pale gold.
    assert banner.font.color.rgb.endswith("10131A")


def test_markup_in_client_data_does_not_break_the_pdf():
    # `Smith & Sons <Holdings>` is a row label in the fixture. reportlab reads
    # a small HTML dialect, so unescaped it raises mid-build and loses the
    # whole report rather than one cell.
    payload, _type, _extension = documents.render_document(_document(), "pdf")
    assert payload.startswith(b"%PDF-")
    assert len(payload) > 3000


def test_a_brand_colour_nobody_validated_cannot_fail_the_job():
    for accent in ("#8a1538", "8A1538", "#abc", "rgb(1,2,3)", "", None, 42, ["#fff"]):
        brand = documents.Brand.for_client(name="Acme", accent=accent)
        assert brand.accent.startswith("#") and len(brand.accent) == 7
        assert brand.accent_ink in ("#ffffff", "#10131a")
        payload, _type, extension = documents.render_document(_document(), "pdf", brand)
        assert extension == "pdf" and payload.startswith(b"%PDF-")


def test_a_long_organisation_name_is_trimmed_rather_than_drawn_off_the_page():
    brand = documents.Brand.for_client(name=" " + "A" * 400 + " ")
    assert len(brand.name) == 60


def test_numbers_are_recovered_from_the_formatting_report_applies():
    assert documents._as_number("£184,320.55") == 184320.55
    assert documents._as_number("(£150.00)") == -150.0
    assert documents._as_number("12,483") == 12483.0
    assert documents._as_number("$99.99") == 99.99
    # Not numbers, and guessing at them would be worse than leaving the text.
    assert documents._as_number("—") is None
    assert documents._as_number("2026-01-01 to 2026-01-31") is None
    assert documents._as_number("{'20': 900}") is None


def test_a_table_with_no_column_names_is_a_definition_list():
    # The provenance trail. With a header row it renders as a bar of brand
    # colour containing nothing at all.
    assert documents._is_definition_table(Table(["", ""], [["Source", "a.xlsx"]]))
    assert not documents._is_definition_table(Table(["Check", "Result"], []))
    assert not documents._is_definition_table(Table(["", "", ""], []))


class _StubSupabase:
    def __init__(self, rows: list[dict]) -> None:
        self.rows = rows
        self.calls: list[str] = []

    def select(self, table: str, **_kwargs: object) -> list[dict]:
        self.calls.append(table)
        return self.rows


class _StubContext:
    def __init__(self, rows: list[dict]) -> None:
        self.supabase = _StubSupabase(rows)
        self.job = {"org_id": "org-1"}


def test_the_report_carries_the_organisation_name_by_default():
    from hermes.jobs import _brand_for

    context = _StubContext([{"id": "org-1", "name": "Acme Accounting"}])
    brand = _brand_for(context, None)

    assert brand.name == "Acme Accounting"
    assert context.supabase.calls == ["organizations"]


def test_a_payload_override_names_the_client_without_a_second_query():
    from hermes.jobs import _brand_for

    context = _StubContext([])
    brand = _brand_for(
        context, {"name": "Kentex Cargo", "accent": "#8a1538", "footer": "Acme · acme.co.uk"}
    )

    assert brand.name == "Kentex Cargo"
    assert brand.accent == "#8a1538"
    assert brand.footer == "Acme · acme.co.uk"
    assert context.supabase.calls == []


def test_an_organisation_that_has_gone_missing_still_produces_a_report():
    from hermes.jobs import _brand_for

    brand = _brand_for(_StubContext([]), {"accent": "#8a1538"})
    assert brand.name == "DataEngine"


def test_characters_the_pdf_font_cannot_draw_are_substituted():
    # `report.py` writes every period comparison with an arrow, and Helvetica
    # has no glyph for it. The `>` of the substitute is then escaped like any
    # other, because escaping is the last thing that happens.
    escaped = documents._escape("£1 → £2")
    assert "-&gt;" in escaped
    assert "£" in escaped, "the pound sign is WinAnsi and must survive"
    assert "&amp;" in documents._escape("Smith & Sons")
    assert "&lt;" in documents._escape("<Holdings>")
