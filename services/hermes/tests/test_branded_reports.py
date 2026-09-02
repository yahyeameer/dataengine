"""
Every format, with a logo and without one.

The point of these is the second half. A report that looks right when the
branding is complete is the easy case; the one that decides whether this feature
is safe to ship is the report generated for an organisation that has set nothing
up, or whose stored logo turns out to be a corrupt file at the moment a client
is waiting for their month-end pack.

So each format is asserted twice: the business name reaches the document either
way, the logo is embedded when there is one, and nothing about a missing or
broken logo can stop the file being produced. The workbook is additionally
checked for the thing decoration usually breaks — that its numbers are still
numbers.
"""

from __future__ import annotations

import datetime as dt
import io
import zipfile

import pytest
from openpyxl import load_workbook

from hermes.tools import branding, documents
from hermes.tools.report import build_report_document, render_markdown

KPIS = {
    "row_count": 12483,
    "period": {"earliest": "2026-09-01", "latest": "2026-09-30"},
    "amount": {"total": 184320.55, "negative_rows": 8},
    "top_by_route": [{"label": "Mombasa - Nairobi", "total": 84200.0, "rows": 410}],
}

SIGNALS = {"exact_duplicates": {"duplicate_rows": 3}, "entity_variants": {"columns": []}}


def _logo(width: int = 240, height: int = 80) -> branding.LogoAsset:
    from PIL import Image

    buffer = io.BytesIO()
    Image.new("RGBA", (width, height), (24, 80, 200, 255)).save(buffer, format="PNG")
    return branding.validate_logo(buffer.getvalue())


def _document():
    return build_report_document(
        workspace_name="Kentex Cargo",
        dataset_name="Monthly Operations Report",
        version_no=4,
        kpis=KPIS,
        profile_signals=SIGNALS,
        generated_at=dt.datetime(2026, 10, 2, 9, 0, tzinfo=dt.timezone.utc),
    )


def _brand(**overrides) -> documents.Brand:
    defaults = {
        "name": "Energy Gain",
        "accent": "#8a1538",
        "footer": "energygain.example · +44 20 7946 0000",
        "period": "September 2026",
    }
    return documents.Brand.for_client(**{**defaults, **overrides})


# ---------------------------------------------------------------------------
# The period, which every format prints
# ---------------------------------------------------------------------------


def test_the_reporting_period_comes_from_the_data_not_the_clock():
    # Rendered in October, about September. Dating it by when it was produced
    # is how a month-end pack gets filed under the wrong month.
    assert _document().period == "September 2026"


def test_a_range_spanning_months_keeps_both_ends():
    document = build_report_document(
        workspace_name="W",
        dataset_name="D",
        version_no=1,
        kpis={"row_count": 1, "period": {"earliest": "2026-01-04", "latest": "2026-03-28"}},
        profile_signals=SIGNALS,
    )
    assert document.period == "January to March 2026"


def test_a_dataset_with_no_dates_prints_no_period():
    document = build_report_document(
        workspace_name="W", dataset_name="D", version_no=1, kpis={"row_count": 1},
        profile_signals=SIGNALS,
    )
    assert document.period == ""


# ---------------------------------------------------------------------------
# PDF (section 15)
# ---------------------------------------------------------------------------


def test_the_pdf_carries_the_business_name_with_a_logo_and_without():
    document = _document()

    with_logo, _type, _ext = documents.render_document(document, "pdf", _brand(logo=_logo()))
    without, _type, _ext = documents.render_document(document, "pdf", _brand())

    for payload in (with_logo, without):
        assert payload.startswith(b"%PDF-")
        assert b"Energy Gain" in payload or b"Energy" in payload

    # The image XObject is the difference between the two, and it is a large
    # one. `/Image` alone appears in every PDF's ProcSet, so the assertion is on
    # the object's own subtype.
    assert len(with_logo) > len(without) + 500
    assert b"/Subtype /Image" in with_logo
    assert b"/Subtype /Image" not in without, "no logo must mean no empty logo box"


def test_a_corrupt_logo_does_not_stop_the_pdf():
    # Section 22. `validate_logo` would have refused these bytes at upload, so
    # reaching the renderer with them means something else went wrong — and the
    # month-end report is not the place to find out.
    broken = branding.LogoAsset(data=b"not a png", mime="image/png", width=100, height=50)
    payload, _type, _ext = documents.render_document(_document(), "pdf", _brand(logo=broken))
    assert payload.startswith(b"%PDF-")
    assert len(payload) > 3000


def test_the_pdf_prints_the_period_and_the_footer():
    payload, _type, _ext = documents.render_document(_document(), "pdf", _brand())
    # reportlab compresses page content, so the text is checked through the
    # extractable strings rather than by scanning the raw bytes.
    text = _pdf_text(payload)
    assert "Energy Gain" in text
    assert "September 2026" in text
    assert "energygain.example" in text


def _pdf_text(payload: bytes) -> str:
    """
    Everything drawn onto the page.

    reportlab writes content streams ASCII85-encoded over Flate, so both layers
    have to come off; a stream that decodes as neither is passed through so a
    change of filter degrades to a weaker assertion rather than a crash.
    """
    import base64
    import re
    import zlib

    chunks: list[str] = []
    for match in re.finditer(rb"stream\r?\n(.*?)endstream", payload, re.S):
        raw = match.group(1).strip()
        for decode in (
            zlib.decompress,
            lambda data: zlib.decompress(base64.a85decode(data, adobe=True)),
        ):
            try:
                chunks.append(decode(raw).decode("latin-1", "ignore"))
                break
            except Exception:  # noqa: BLE001 - an image stream is not text
                continue
    return "\n".join(chunks)


# ---------------------------------------------------------------------------
# DOCX (section 16)
# ---------------------------------------------------------------------------


def test_the_word_file_embeds_a_real_picture_a_recipient_can_replace():
    payload, _type, _ext = documents.render_document(_document(), "docx", _brand(logo=_logo()))

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        names = archive.namelist()
        media = [name for name in names if name.startswith("word/media/")]
        document_xml = archive.read("word/document.xml").decode("utf-8")

    # A real drawing in the document body, not a picture of the whole header.
    assert len(media) == 1
    assert "<w:drawing>" in document_xml
    assert "Energy Gain" in document_xml
    assert "September 2026" in document_xml


def test_the_word_file_without_a_logo_has_no_media_and_still_names_the_business():
    payload, _type, _ext = documents.render_document(_document(), "docx", _brand())

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert [name for name in archive.namelist() if name.startswith("word/media/")] == []
        assert "Energy Gain" in archive.read("word/document.xml").decode("utf-8")


def test_a_corrupt_logo_does_not_stop_the_word_file():
    broken = branding.LogoAsset(data=b"not a png", mime="image/png", width=100, height=50)
    payload, _type, _ext = documents.render_document(_document(), "docx", _brand(logo=broken))
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert archive.testzip() is None
        assert "Energy Gain" in archive.read("word/document.xml").decode("utf-8")


# ---------------------------------------------------------------------------
# XLSX (section 17)
# ---------------------------------------------------------------------------


def test_the_workbook_is_branded_without_losing_its_numbers():
    payload, _type, _ext = documents.render_document(_document(), "xlsx", _brand(logo=_logo()))
    sheet = load_workbook(io.BytesIO(payload)).active

    assert sheet.cell(row=2, column=2).value == "Energy Gain"
    assert sheet.cell(row=2, column=6).value == "September 2026"

    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert 184320.55 in values, "the headline total must stay a number Excel can sum"
    assert 12483 in values or 12483.0 in values


def test_the_workbook_carries_the_logo_as_a_drawing():
    payload, _type, _ext = documents.render_document(_document(), "xlsx", _brand(logo=_logo()))
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert any(name.startswith("xl/media/") for name in archive.namelist())

    plain, _type, _ext = documents.render_document(_document(), "xlsx", _brand())
    with zipfile.ZipFile(io.BytesIO(plain)) as archive:
        assert not any(name.startswith("xl/media/") for name in archive.namelist())


def test_the_accent_reaches_the_workbook_band():
    payload, _type, _ext = documents.render_document(_document(), "xlsx", _brand())
    sheet = load_workbook(io.BytesIO(payload)).active
    assert sheet.cell(row=2, column=2).fill.start_color.rgb.endswith("8A1538")
    # White ink on a dark maroon, chosen by the same rule the settings preview
    # uses rather than assumed.
    assert sheet.cell(row=2, column=2).font.color.rgb.endswith("FFFFFF")


# ---------------------------------------------------------------------------
# Markdown (section 18)
# ---------------------------------------------------------------------------


def test_markdown_leads_with_the_business_name_and_then_the_report():
    text = render_markdown(_document(), _brand())
    assert text.startswith("# Energy Gain\n")
    assert "## Monthly Operations Report" in text
    assert "September 2026" in text
    assert "energygain.example" in text


def test_markdown_shows_no_image_for_a_private_logo():
    # The private logo is embedded in the three binary formats and cannot be in
    # this one, so the header is text. No signed URL is ever printed here.
    text = render_markdown(_document(), _brand(logo=_logo()))
    assert "![" not in text


def test_markdown_uses_a_public_logo_url_when_an_administrator_supplied_one():
    text = render_markdown(_document(), _brand(logo_url="https://cdn.example.com/logo.png"))
    assert "![Energy Gain](https://cdn.example.com/logo.png)" in text


def test_markdown_without_a_brand_is_unchanged():
    # The assistant renders this in the thread for reports that have no resolved
    # organisation identity, and printing a placeholder owner would be a claim.
    text = render_markdown(_document())
    assert text.startswith("# Monthly Operations Report\n")


# ---------------------------------------------------------------------------
# Several formats at once (section 22)
# ---------------------------------------------------------------------------


def test_render_all_produces_every_requested_format():
    results = documents.render_all(_document(), ["pdf", "xlsx", "md"], _brand())
    assert [item.format for item in results] == ["pdf", "xlsx", "md"]
    assert all(item.ok for item in results)


def test_one_failing_renderer_does_not_take_the_others_with_it():
    document = _document()

    def explode(*_args, **_kwargs):
        raise RuntimeError("the chart engine fell over")

    original = documents.render_xlsx
    documents.render_xlsx = explode  # type: ignore[assignment]
    try:
        results = documents.render_all(document, ["pdf", "xlsx"], _brand())
    finally:
        documents.render_xlsx = original  # type: ignore[assignment]

    by_format = {item.format: item for item in results}
    assert by_format["pdf"].ok, "a format that rendered must survive another one failing"
    assert not by_format["xlsx"].ok
    assert "chart engine" in (by_format["xlsx"].error or "")


def test_an_unknown_format_is_reported_rather_than_silently_dropped():
    results = documents.render_all(_document(), ["pptx"], _brand())
    assert results[0].ok is False
    assert "not a report format" in (results[0].error or "")


# ---------------------------------------------------------------------------
# Branding that is missing entirely
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("fmt", documents.FORMATS)
def test_no_branding_at_all_still_produces_every_format(fmt):
    resolved = branding.resolve_branding()
    brand = documents.Brand.for_client(
        name=resolved.business_name, accent=resolved.accent, footer=resolved.footer
    )
    payload, _type, extension = documents.render_document(_document(), fmt, brand)
    assert extension == fmt
    assert len(payload) > 500
