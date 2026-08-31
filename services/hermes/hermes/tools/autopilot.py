"""
The pieces of the one-click categorisation run that are worth testing alone.

Column choice and export validation both live here rather than in `jobs.py` for
the same reason: they are pure functions over data, they are where this feature
is most likely to be wrong, and a test for either of them should not have to
stand up a fake Supabase to run.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any

from . import hmrc

# What a bank calls the column holding the merchant, roughly in the order a
# statement is likely to use them.
#
# Ordered rather than scored. A statement can carry both `Type` (DD, POS, TFR)
# and `Transaction`, and only the second one names anybody -- so the list has to
# express a preference, not a set.
_DESCRIPTION_NAMES = (
    "transaction",
    "description",
    "details",
    "narrative",
    "payee",
    "merchant",
    "vendor",
    "supplier",
    "counterparty",
    "particulars",
    "memo",
    "reference",
    "type",
)

_DATE_NAMES = ("date", "transaction date", "posted", "value date")
_AMOUNT_NAMES = ("amount", "value", "out", "in", "debit", "credit", "paid out", "paid in")

# How much of a column the HMRC rules must recognise before it counts as the
# description on evidence alone. Low on purpose: a statement for a niche trade
# can be mostly merchants we have never seen, and one line in seven matching is
# still far more than any other column in the file will manage.
_MIN_HIT_RATE = 0.15


def _rank(name: str, candidates: tuple[str, ...]) -> int:
    """Position in the preference list, or a large number for no match."""
    lowered = name.strip().lower()
    for index, candidate in enumerate(candidates):
        if lowered == candidate:
            return index
        if candidate in lowered:
            # A containment match ranks below every exact match, so
            # "Transaction Date" cannot outrank a column literally called
            # "Description".
            return index + len(candidates)
    return len(candidates) * 3


def choose_description_column(columns: list[str], rows: list[dict[str, Any]]) -> str | None:
    """
    The column a UK bank statement puts the merchant in.

    Two signals, and the second is what stops this being a guessing game. The
    name gets a column onto the shortlist; whether the HMRC rules actually
    recognise anything in it decides the winner. A file whose columns are called
    `Col1..Col6` -- which is what a PDF-to-Excel conversion produces -- has no
    usable names at all, and the only honest way to find the description is to
    look at the values.

    Returns None when nothing looks like a description, which the caller reports
    as "we could not find the transactions in this file" rather than
    categorising the date column.
    """
    usable = [name for name in columns if name and not name.startswith("__")]
    if not usable:
        return None

    sample = rows[:400]
    scored: list[tuple[float, float, int, str]] = []

    for name in usable:
        values = [row.get(name) for row in sample]
        text_values = [value for value in values if isinstance(value, str) and value.strip()]
        if not text_values:
            # A column of numbers is an amount, not a description, whatever it
            # is called.
            continue

        matched = sum(1 for value in text_values if hmrc.categorise(value) is not None)
        hit_rate = matched / len(text_values)

        # Distinctness separates a description from a category-like column: a
        # `Type` column has five distinct values in four hundred rows. It is a
        # tie-breaker only -- on its own it would elect the date column, which
        # is text, perfectly distinct, and not a description of anything.
        distinct = len({value.strip().lower() for value in text_values})
        variety = min(distinct / max(len(text_values), 1) * 4, 1.0)

        scored.append((hit_rate, variety, _rank(name, _DESCRIPTION_NAMES), name))

    if not scored:
        return None

    # Two tiers, and the order between them is the whole heuristic.
    #
    # Evidence first: a column the HMRC rules recognise *is* the description,
    # whatever it is called, which is what makes this work on a file whose
    # headers are `Col1..Col6`. Only when nothing is recognised -- a statement
    # full of merchants we have never seen -- does the column name get to decide,
    # and if neither says anything we refuse.
    recognised = [entry for entry in scored if entry[0] >= _MIN_HIT_RATE]
    if recognised:
        recognised.sort(key=lambda entry: (-entry[0], -entry[1], entry[2]))
        return recognised[0][3]

    named = [entry for entry in scored if entry[2] < len(_DESCRIPTION_NAMES) * 2]
    if named:
        named.sort(key=lambda entry: (entry[2], -entry[1]))
        return named[0][3]

    # Nothing recognised and nothing named like a description. Refusing beats
    # picking the first text column and handing back a file whose every row
    # reads 'Uncategorised' -- that looks like the product failing rather than
    # like the file being the wrong shape.
    return None


def find_column(columns: list[str], candidates: tuple[str, ...]) -> str | None:
    """The first column matching a preference list, or None."""
    ranked = sorted(
        ((_rank(name, candidates), index, name) for index, name in enumerate(columns)),
        key=lambda item: (item[0], item[1]),
    )
    for rank, _index, name in ranked:
        if rank < len(candidates) * 3:
            return name
    return None


# -----------------------------------------------------------------------------
# Validating the file before anybody is told it is ready
# -----------------------------------------------------------------------------


class ValidationError(RuntimeError):
    """The produced workbook is not fit to hand over."""


@dataclass
class ExportCheck:
    rows: int
    columns: list[str]
    categorised: int
    flagged: int


def validate_export(
    workbook_bytes: bytes,
    *,
    expected_rows: int,
    source_column: str,
    source_values: list[Any],
) -> ExportCheck:
    """
    Open the workbook that was just written and prove it is the right file.

    This is the difference between "the model returned results" and "the
    accountant has a categorised file", and the brief is blunt about which one
    counts. Everything up to this point is an intention: a mapping was approved,
    an operation reported success, bytes were uploaded. None of that has opened
    the file.

    So the file is opened, with the same library anybody else would use, and
    seven things are checked:

      1. it parses at all
      2. it has a header row and the expected number of data rows
      3. the description column survived, value for value
      4. the three HMRC columns exist
      5. every row has a category -- no blanks
      6. every category is one this taxonomy declares
      7. every non-empty box is a real SA103F box for its category

    Any failure raises, and the job fails. A wrong file that says it is right is
    worse than no file, because the accountant has no reason to check it.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises a wide variety here
        raise ValidationError(f"the workbook could not be opened: {error}") from error

    try:
        sheet = workbook.active
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not grid:
        raise ValidationError("the workbook is empty")

    header = [str(cell) if cell is not None else "" for cell in grid[0]]
    body = grid[1:]

    if len(body) != expected_rows:
        raise ValidationError(
            f"the workbook holds {len(body)} row(s) but the version has {expected_rows}"
        )

    missing = [name for name in hmrc.OUTPUT_COLUMNS if name not in header]
    if missing:
        raise ValidationError(f"the workbook is missing {', '.join(missing)}")

    if source_column not in header:
        raise ValidationError(
            f"the original column {source_column!r} is not in the workbook"
        )

    index_of = {name: position for position, name in enumerate(header)}
    source_index = index_of[source_column]
    category_index = index_of[hmrc.CATEGORY_COLUMN]
    box_index = index_of[hmrc.BOX_COLUMN]
    confidence_index = index_of[hmrc.CONFIDENCE_COLUMN]

    categorised = 0
    flagged = 0
    needs_review_label = hmrc.CONFIDENCE_LABELS[hmrc.LOW]

    for position, row in enumerate(body):
        original = row[source_index] if source_index < len(row) else None
        expected = source_values[position] if position < len(source_values) else None
        # Compared as text: the workbook has been through openpyxl's type
        # handling and a float that arrived as a string is not a corruption.
        if _as_text(original) != _as_text(expected):
            raise ValidationError(
                f"row {position + 2} lost its original {source_column!r} value"
            )

        category = _as_text(row[category_index] if category_index < len(row) else None)
        if not category:
            raise ValidationError(f"row {position + 2} has no category")
        if category not in hmrc.CATEGORY_NAMES:
            raise ValidationError(f"row {position + 2} has an unknown category {category!r}")

        box = _as_text(row[box_index] if box_index < len(row) else None)
        expected_box = hmrc.BOX_BY_CATEGORY.get(category, "")
        if box != expected_box:
            raise ValidationError(
                f"row {position + 2} maps {category!r} to {box or 'no box'}, "
                f"which is not its box"
            )

        confidence = _as_text(row[confidence_index] if confidence_index < len(row) else None)
        if confidence == needs_review_label or category == hmrc.FALLBACK:
            flagged += 1
        else:
            categorised += 1

    return ExportCheck(
        rows=len(body), columns=header, categorised=categorised, flagged=flagged
    )


def _as_text(value: Any) -> str:
    """
    A cell as comparable text.

    Numbers survive a round trip through xlsx as floats, so 12.5 and "12.5" have
    to compare equal or every amount column would look corrupted. Trailing ".0"
    on a whole number is the same story.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)
