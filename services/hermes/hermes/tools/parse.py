"""
Messy workbook parsing (PRD section 6, a P0 deliverable rather than an
assumption -- "this is where pilots die").

The input is a grid of cells that a human maintained. It has a title block, a
blank row, a header somewhere around row 5, data with blank separators and an
embedded subtotal, a trailing TOTAL that must never be treated as a
transaction, footnotes underneath, and a second sheet holding a different table
entirely.

The job is to find the table and say how sure it is. Everything downstream --
profiling, proposals, the recipe -- is built on this answer, so the parser
reports a confidence and its reasoning rather than returning a DataFrame and
hoping.

No LLM is involved. Structure detection is deterministic and testable, and the
eval harness in section 8 only means something if the thing it measures does
the same thing twice.
"""

from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
from collections import Counter
from dataclasses import asdict, dataclass, field
from typing import Any, Literal

from .values import (
    entity_key,
    is_null_token,
    is_subtotal_label,
    normalize_text,
    parse_date,
    parse_number,
)

log = logging.getLogger("hermes.parse")

ColumnType = Literal["date", "number", "text", "boolean", "empty"]

# How far down to look for a header before concluding the sheet has none. A
# title block of more than this is a document, not a data export.
MAX_HEADER_SEARCH_ROWS = 40

# Consecutive blank rows that end a table. Two is too few -- the fixture has a
# blank row inside its data, and so does every hand-maintained ledger.
BLANK_RUN_ENDS_TABLE = 3


@dataclass
class SkippedRow:
    """A row excluded from the data, and why. Section 7 needs the why."""

    source_row: int
    reason: Literal["blank", "subtotal", "footnote", "above_header", "sparse"]
    preview: str


@dataclass
class ColumnInterpretation:
    index: int
    source_header: str
    name: str
    inferred_type: ColumnType
    type_confidence: float
    non_null: int
    parse_failures: int
    date_order: str | None = None
    number_styles: list[str] = field(default_factory=list)
    failure_samples: list[str] = field(default_factory=list)
    ambiguous_dates: int = 0


@dataclass
class SheetInterpretation:
    sheet_name: str
    header_row: int | None
    first_data_row: int | None
    last_data_row: int | None
    first_column: int
    last_column: int
    data_rows: int
    columns: list[ColumnInterpretation]
    skipped: list[SkippedRow]
    confidence: float
    notes: list[str]

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "first_data_row": self.first_data_row,
            "last_data_row": self.last_data_row,
            "first_column": self.first_column,
            "last_column": self.last_column,
            "data_rows": self.data_rows,
            "columns": [asdict(column) for column in self.columns],
            "skipped": [asdict(row) for row in self.skipped],
            "confidence": round(self.confidence, 3),
            "notes": self.notes,
        }


@dataclass
class ParsedTable:
    """One recognised table: its interpretation plus the values behind it."""

    interpretation: SheetInterpretation
    # Column-major, coerced. Parallel to interpretation.columns.
    columns: dict[str, list[Any]]
    # The original row number in the sheet for each output row. This is what
    # makes "show me the source rows" (section 7) a lookup rather than a guess.
    source_rows: list[int]

    @property
    def row_count(self) -> int:
        return len(self.source_rows)


@dataclass
class ParseResult:
    tables: list[ParsedTable]
    primary_index: int
    source_signature: str
    warnings: list[str]

    @property
    def primary(self) -> ParsedTable:
        return self.tables[self.primary_index]

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_signature": self.source_signature,
            "primary_sheet": self.tables[self.primary_index].interpretation.sheet_name,
            "sheets": [table.interpretation.to_dict() for table in self.tables],
            "warnings": self.warnings,
        }


# -----------------------------------------------------------------------------
# Reading bytes into grids
# -----------------------------------------------------------------------------


def _read_xlsx(data: bytes) -> list[tuple[str, list[list[Any]]]]:
    from openpyxl import load_workbook

    # data_only=True returns the cached result of a formula rather than the
    # formula text. An accountant's file is full of formulas and the value is
    # what they see on screen; the formula would parse as a string and poison
    # every type inference in the column.
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    sheets: list[tuple[str, list[list[Any]]]] = []

    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            # A hidden sheet is usually a lookup table or a previous month left
            # behind. Parsing it as the main table is a classic wrong answer.
            continue

        grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
        _fill_merged_cells(worksheet, grid)
        sheets.append((worksheet.title, grid))

    workbook.close()
    return sheets


def _fill_merged_cells(worksheet: Any, grid: list[list[Any]]) -> None:
    """
    Propagate a merged range's value across the cells it covers.

    openpyxl returns the value in the top-left cell and None everywhere else,
    which makes a merged header look like a header followed by blanks -- and a
    merged title block look like a populated row, which is exactly what the
    header scorer must not be fooled by.
    """
    for merged in list(worksheet.merged_cells.ranges):
        min_row, min_col = merged.min_row, merged.min_col
        if min_row - 1 >= len(grid):
            continue
        row = grid[min_row - 1]
        if min_col - 1 >= len(row):
            continue

        value = row[min_col - 1]
        if value is None:
            continue

        for r in range(merged.min_row - 1, min(merged.max_row, len(grid))):
            for c in range(merged.min_col - 1, merged.max_col):
                if c < len(grid[r]):
                    grid[r][c] = value


def _read_csv(data: bytes) -> list[tuple[str, list[list[Any]]]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Sniffer fails on a single-column file or one with quoted commas only.
        # Counting is cruder and more reliable.
        counts = {sep: sample.count(sep) for sep in ",;\t|"}
        delimiter = max(counts, key=lambda key: counts[key]) if any(counts.values()) else ","

    rows = [list(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return [("Sheet1", rows)]


def read_grids(data: bytes, filename: str) -> list[tuple[str, list[list[Any]]]]:
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_xlsx(data)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return _read_csv(data)
    if lower.endswith(".xls"):
        # The legacy binary format needs a different reader entirely. Refusing
        # clearly beats mis-parsing: the upload path already accepts .xls, so an
        # accountant will hit this, and "convert to .xlsx" is an instruction
        # they can act on in ten seconds.
        raise ValueError(
            "Legacy .xls files are not supported yet. Re-save the workbook as .xlsx and upload again."
        )
    raise ValueError(f"Cannot parse {filename!r}: expected .xlsx, .xlsm or .csv")


# -----------------------------------------------------------------------------
# Structure detection
# -----------------------------------------------------------------------------


def _cell_populated(value: Any) -> bool:
    return not is_null_token(value)


def _row_populated_count(row: list[Any]) -> int:
    return sum(1 for cell in row if _cell_populated(cell))


def _looks_like_header_cell(value: Any) -> bool:
    """
    A header is short text that is not a number and not a date.

    The length ceiling matters: a footnote or a disclaimer sentence is text too,
    and without a limit a 90-character legal notice scores as a perfect header.
    """
    if not _cell_populated(value):
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return False

    text = normalize_text(value)
    if len(text) > 60:
        return False
    if parse_number(text).ok:
        return False
    if parse_date(text).ok:
        return False
    return bool(re.search(r"[A-Za-z]", text))


def _cell_shape(value: Any) -> str | None:
    """A coarse description of what a cell is, for comparing like with like."""
    if not _cell_populated(value):
        return None
    if parse_number(value).ok:
        return "number"
    if parse_date(value).ok:
        return "date"
    return "short_text" if len(normalize_text(value)) <= 30 else "long_text"


def _column_contrast(grid: list[list[Any]], index: int) -> float:
    """
    How unlike its own column each cell of a candidate row is.

    This is the signal that survives when the length ceiling fails. A header
    does not resemble the values underneath it -- "Timestamp" sits above
    timestamps, a question sits above answers -- whereas a data row is
    indistinguishable from its neighbours by construction, because it *is* one
    of them.

    It matters most for the file that exposed the gap: a survey export whose
    headers are the questions, twenty-five of them past sixty characters and so
    rejected as labels, above short human answers that read like perfect
    headers. Judged on length alone the nineteenth respondent won, her answers
    became the column names, and the eighteen people above her were discarded as
    if they sat above the table. Judged on contrast she loses, because her
    answers look exactly like everyone else's.

    Deliberately coarse. Comparing shapes -- number, date, short text, long text
    -- rather than values keeps it robust on a column whose entries are all
    different, which is most of them.
    """
    row = grid[index]
    body = grid[index + 1 : index + 21]
    if not body:
        return 0.0

    considered = 0
    differing = 0

    for column, cell in enumerate(row):
        shape = _cell_shape(cell)
        if shape is None:
            continue

        below = [
            _cell_shape(body_row[column])
            for body_row in body
            if column < len(body_row) and _cell_populated(body_row[column])
        ]
        if len(below) < 3:
            continue

        considered += 1
        modal = Counter(below).most_common(1)[0][0]
        if shape != modal:
            differing += 1

    return (differing / considered) if considered else 0.0


def _score_header_row(grid: list[list[Any]], index: int, body_width: int) -> tuple[float, list[str]]:
    """
    Score a candidate header row.

    The signal is not "this row contains text". It is "this row contains text
    *and the rows beneath it do not*". A title block is text with nothing
    structured under it; a header is text sitting on top of a typed table. The
    contrast between the row and its body is what separates row 5 from rows 1-3
    in the fixture.
    """
    row = grid[index]
    populated = [cell for cell in row if _cell_populated(cell)]
    if len(populated) < 2:
        return 0.0, ["fewer than two populated cells"]

    reasons: list[str] = []
    header_like = sum(1 for cell in row if _looks_like_header_cell(cell))
    header_ratio = header_like / len(populated)

    labels = [normalize_text(cell).lower() for cell in populated]
    distinct_ratio = len(set(labels)) / len(labels)

    body = grid[index + 1 : index + 11]
    body_populated = [_row_populated_count(r) for r in body]
    non_empty_body = [count for count in body_populated if count > 0]
    if not non_empty_body:
        return 0.0, ["nothing beneath it"]

    # Width agreement: a header spans the same columns its data does.
    width_match = 1.0 - min(1.0, abs(len(populated) - body_width) / max(body_width, 1))

    # Typed contrast: the body should be more numeric/date-ish than the header.
    body_typed = 0
    body_cells = 0
    for body_row in body:
        for cell in body_row:
            if not _cell_populated(cell):
                continue
            body_cells += 1
            if parse_number(cell).ok or parse_date(cell).ok:
                body_typed += 1
    typed_contrast = (body_typed / body_cells) if body_cells else 0.0

    # How unlike its own column this row is. Carries real weight because it is
    # the only term that still works when the labels are long: header_ratio
    # measures whether cells look like labels, and a survey question does not.
    column_contrast = _column_contrast(grid, index)

    score = (
        0.25 * header_ratio
        + 0.15 * distinct_ratio
        + 0.20 * width_match
        + 0.15 * min(1.0, typed_contrast * 1.5)
        + 0.25 * column_contrast
    )

    if column_contrast > 0.6:
        reasons.append("does not resemble the values in its own columns")
    elif column_contrast < 0.2:
        reasons.append("looks like one of the rows beneath it")

    # A row of text above a row of text is a title block, not a header.
    #
    # Guarded by column_contrast now: a genuine header of long questions also
    # has no typed values beneath it when every answer is text, and demoting it
    # here is what let a data row win outright.
    if typed_contrast < 0.05 and header_ratio > 0.9 and column_contrast < 0.5:
        score *= 0.4
        reasons.append("no typed values beneath; may be a title block")

    if header_ratio == 1.0 and distinct_ratio == 1.0 and width_match > 0.9:
        reasons.append("all cells are distinct short labels spanning the data width")

    return score, reasons


def _modal_body_width(grid: list[list[Any]], start: int) -> int:
    """The width most rows agree on — the table's real column count."""
    counts = Counter(
        _row_populated_count(row) for row in grid[start : start + 60] if _row_populated_count(row) > 1
    )
    return counts.most_common(1)[0][0] if counts else 0


def find_header_row(grid: list[list[Any]]) -> tuple[int | None, float, list[str]]:
    limit = min(len(grid) - 1, MAX_HEADER_SEARCH_ROWS)
    best_index: int | None = None
    best_score = 0.0
    best_reasons: list[str] = []

    for index in range(limit):
        if _row_populated_count(grid[index]) < 2:
            continue
        body_width = _modal_body_width(grid, index + 1)
        if body_width < 2:
            continue

        score, reasons = _score_header_row(grid, index, body_width)
        if score > best_score:
            best_index, best_score, best_reasons = index, score, reasons

    return best_index, best_score, best_reasons


def _column_bounds(grid: list[list[Any]], header_index: int) -> tuple[int, int]:
    header = grid[header_index]
    populated = [i for i, cell in enumerate(header) if _cell_populated(cell)]
    if not populated:
        return 0, 0
    return populated[0], populated[-1]


def _classify_row(
    row: list[Any], first_col: int, last_col: int, width: int
) -> Literal["data", "blank", "subtotal", "sparse"]:
    window = row[first_col : last_col + 1]
    populated = _row_populated_count(window)

    if populated == 0:
        return "blank"

    # A summary label anywhere in the leading cells marks the row as a subtotal.
    # Checking only the first cell misses "  Subtotal" indented into column B,
    # which is how most people format them.
    for cell in window[: min(3, len(window))]:
        if is_subtotal_label(cell):
            return "subtotal"

    # A row with one populated cell under a five-column header is a footnote or
    # a section heading, not a transaction.
    if populated == 1 and width >= 3:
        return "sparse"

    return "data"


def detect_table(grid: list[list[Any]], sheet_name: str) -> tuple[SheetInterpretation, list[int]]:
    """Locate the table and return its interpretation plus the data row indices."""
    notes: list[str] = []
    skipped: list[SkippedRow] = []

    header_index, header_score, header_reasons = find_header_row(grid)
    notes.extend(header_reasons)

    if header_index is None:
        return (
            SheetInterpretation(
                sheet_name=sheet_name,
                header_row=None,
                first_data_row=None,
                last_data_row=None,
                first_column=0,
                last_column=0,
                data_rows=0,
                columns=[],
                skipped=[],
                confidence=0.0,
                notes=notes + ["no header row found"],
            ),
            [],
        )

    if header_index > 0:
        notes.append(f"header is on row {header_index + 1}, not row 1")
        for index in range(header_index):
            if _row_populated_count(grid[index]) > 0:
                skipped.append(
                    SkippedRow(index + 1, "above_header", _preview(grid[index]))
                )

    first_col, last_col = _column_bounds(grid, header_index)
    width = last_col - first_col + 1

    data_indices: list[int] = []
    blank_run = 0
    table_ended = False

    for index in range(header_index + 1, len(grid)):
        row = grid[index]
        kind = _classify_row(row, first_col, last_col, width)

        if table_ended:
            if _row_populated_count(row) > 0:
                skipped.append(SkippedRow(index + 1, "footnote", _preview(row)))
            continue

        if kind == "blank":
            blank_run += 1
            if blank_run >= BLANK_RUN_ENDS_TABLE:
                table_ended = True
            continue

        blank_run = 0

        if kind == "subtotal":
            # Not an error and not data. Kept as evidence: a subtotal is a
            # checksum the accountant already computed, and section 5.3's
            # post-run invariants can reconcile the cleaned total against it.
            skipped.append(SkippedRow(index + 1, "subtotal", _preview(row)))
            continue

        if kind == "sparse":
            skipped.append(SkippedRow(index + 1, "footnote", _preview(row)))
            continue

        data_indices.append(index)

    # Trailing summary rows sometimes read as data because they fill enough
    # columns. Trim from the bottom while the row looks like a total.
    while data_indices:
        last = data_indices[-1]
        window = grid[last][first_col : last_col + 1]
        if any(is_subtotal_label(cell) for cell in window):
            skipped.append(SkippedRow(last + 1, "subtotal", _preview(grid[last])))
            data_indices.pop()
            continue
        break

    headers = [normalize_text(grid[header_index][col]) if col < len(grid[header_index]) else ""
               for col in range(first_col, last_col + 1)]

    columns = _infer_columns(grid, data_indices, headers, first_col, last_col)

    confidence = _overall_confidence(header_score, columns, len(data_indices))

    if any(row.reason == "subtotal" for row in skipped):
        notes.append(
            f"{sum(1 for r in skipped if r.reason == 'subtotal')} summary row(s) excluded from the data"
        )
    if any(row.reason == "footnote" for row in skipped):
        notes.append(
            f"{sum(1 for r in skipped if r.reason == 'footnote')} footnote or heading row(s) excluded"
        )

    interpretation = SheetInterpretation(
        sheet_name=sheet_name,
        header_row=header_index + 1,
        first_data_row=(data_indices[0] + 1) if data_indices else None,
        last_data_row=(data_indices[-1] + 1) if data_indices else None,
        first_column=first_col + 1,
        last_column=last_col + 1,
        data_rows=len(data_indices),
        columns=columns,
        skipped=skipped,
        confidence=confidence,
        notes=notes,
    )
    return interpretation, data_indices


def _preview(row: list[Any], limit: int = 80) -> str:
    parts = [normalize_text(cell) for cell in row if _cell_populated(cell)]
    text = " | ".join(parts)
    return text[:limit] + ("…" if len(text) > limit else "")


def _column_name(header: str, index: int, taken: set[str]) -> str:
    """
    A stable machine name for a column.

    Derived from the header rather than assigned positionally, because a recipe
    written in month 1 has to survive a month-2 file where someone inserted a
    column -- matching on `net_sales` survives that, matching on `column_4`
    does not.
    """
    base = re.sub(r"[^\w]+", "_", normalize_text(header).lower()).strip("_")
    base = re.sub(r"_{2,}", "_", base)
    if not base or base[0].isdigit():
        base = f"col_{index + 1}" if not base else f"c_{base}"

    name = base
    suffix = 2
    while name in taken:
        name = f"{base}_{suffix}"
        suffix += 1
    taken.add(name)
    return name


def _infer_columns(
    grid: list[list[Any]],
    data_indices: list[int],
    headers: list[str],
    first_col: int,
    last_col: int,
) -> list[ColumnInterpretation]:
    columns: list[ColumnInterpretation] = []
    taken: set[str] = set()

    for offset, col in enumerate(range(first_col, last_col + 1)):
        raw_values = [
            grid[index][col] if col < len(grid[index]) else None for index in data_indices
        ]
        non_null = [value for value in raw_values if _cell_populated(value)]

        header = headers[offset] if offset < len(headers) else ""
        name = _column_name(header, offset, taken)

        if not non_null:
            columns.append(
                ColumnInterpretation(
                    index=offset,
                    source_header=header,
                    name=name,
                    inferred_type="empty",
                    type_confidence=1.0,
                    non_null=0,
                    parse_failures=0,
                )
            )
            continue

        number_hits = 0
        number_styles: Counter[str] = Counter()
        date_hits = 0
        date_orders: Counter[str] = Counter()
        ambiguous_dates = 0
        bool_hits = 0

        for value in non_null:
            parsed_number = parse_number(value)
            if parsed_number.ok:
                number_hits += 1
                number_styles.update(parsed_number.styles)

            parsed_date = parse_date(value)
            if parsed_date.ok:
                date_hits += 1
                date_orders[parsed_date.order] += 1
                if parsed_date.ambiguous:
                    ambiguous_dates += 1

            text = normalize_text(value).lower()
            if text in {"true", "false", "yes", "no", "y", "n"}:
                bool_hits += 1

        total = len(non_null)
        number_ratio = number_hits / total
        date_ratio = date_hits / total
        bool_ratio = bool_hits / total

        # Dates win ties against numbers. A four-digit year parses as a number
        # and an Excel serial parses as both, so preferring "number" would turn
        # every date column into a numeric one.
        if date_ratio >= 0.7 and date_ratio >= number_ratio:
            inferred: ColumnType = "date"
            confidence = date_ratio
            failures = [v for v in non_null if not parse_date(v).ok]
        elif number_ratio >= 0.7:
            inferred = "number"
            confidence = number_ratio
            failures = [v for v in non_null if not parse_number(v).ok]
        elif bool_ratio >= 0.9:
            inferred = "boolean"
            confidence = bool_ratio
            failures = []
        else:
            inferred = "text"
            # Text is the fallback, so its confidence is "how clearly is this
            # *not* something else" rather than a positive signal.
            confidence = 1.0 - max(number_ratio, date_ratio)
            failures = []

        # The column-level date decision. Rows where one component exceeds 12
        # prove the convention; ambiguous rows then follow that proof instead of
        # each guessing on its own.
        date_order = None
        if inferred == "date":
            proven = {order: count for order, count in date_orders.items() if order in {"dmy", "mdy"}}
            unambiguous = [
                parse_date(value).order
                for value in non_null
                if parse_date(value).ok and not parse_date(value).ambiguous
            ]
            proven_orders = Counter(o for o in unambiguous if o in {"dmy", "mdy"})
            if proven_orders:
                date_order = proven_orders.most_common(1)[0][0]
            elif proven:
                date_order = max(proven, key=lambda key: proven[key])
            else:
                date_order = "dmy"

        columns.append(
            ColumnInterpretation(
                index=offset,
                source_header=header,
                name=name,
                inferred_type=inferred,
                type_confidence=round(confidence, 3),
                non_null=total,
                parse_failures=len(failures),
                date_order=date_order,
                number_styles=sorted(number_styles),
                failure_samples=[normalize_text(v)[:40] for v in failures[:5]],
                ambiguous_dates=ambiguous_dates,
            )
        )

    return columns


def _overall_confidence(
    header_score: float, columns: list[ColumnInterpretation], data_rows: int
) -> float:
    if not columns or data_rows == 0:
        return 0.0

    typed = [c for c in columns if c.inferred_type in {"date", "number", "boolean"}]
    type_confidence = (
        sum(c.type_confidence for c in typed) / len(typed) if typed else 0.5
    )
    # A two-row table can look perfect and mean nothing. Scale by volume so the
    # confidence reported to the accountant reflects the evidence behind it.
    volume = min(1.0, data_rows / 20)

    return round(0.45 * header_score + 0.40 * type_confidence + 0.15 * volume, 3)


# -----------------------------------------------------------------------------
# Coercion into columnar data
# -----------------------------------------------------------------------------


def build_table(
    grid: list[list[Any]], interpretation: SheetInterpretation, data_indices: list[int]
) -> ParsedTable:
    """
    Materialise the detected table, applying the column-level type decisions.

    A value that will not coerce becomes None and is already counted in
    `parse_failures` -- it is not dropped and it is not guessed at. The raw text
    stays available in the `__raw_` companion column so the cleaned version can
    still answer "what did the file actually say here" without re-reading the
    original workbook.
    """
    first_col = interpretation.first_column - 1
    columns: dict[str, list[Any]] = {}
    source_rows = [index + 1 for index in data_indices]

    for column in interpretation.columns:
        col_index = first_col + column.index
        values: list[Any] = []
        raw_values: list[str | None] = []

        for row_index in data_indices:
            row = grid[row_index]
            cell = row[col_index] if col_index < len(row) else None
            raw_values.append(None if not _cell_populated(cell) else normalize_text(cell))

            if not _cell_populated(cell):
                values.append(None)
                continue

            if column.inferred_type == "number":
                parsed = parse_number(cell)
                values.append(parsed.as_float if parsed.ok else None)
            elif column.inferred_type == "date":
                order = column.date_order if column.date_order in {"dmy", "mdy"} else "dmy"
                parsed_date = parse_date(cell, prefer=order)  # type: ignore[arg-type]
                values.append(parsed_date.value.isoformat() if parsed_date.ok and parsed_date.value else None)
            elif column.inferred_type == "boolean":
                text = normalize_text(cell).lower()
                values.append(text in {"true", "yes", "y", "1"})
            else:
                values.append(normalize_text(cell))

        columns[column.name] = values
        # Only kept where coercion actually changed something. Carrying a raw
        # copy of every text column would double the Parquet for no benefit.
        if column.inferred_type in {"number", "date"}:
            columns[f"__raw_{column.name}"] = raw_values

    return ParsedTable(interpretation=interpretation, columns=columns, source_rows=source_rows)


def compute_source_signature(tables: list[ParsedTable]) -> str:
    """
    A fingerprint of the file's shape, used to auto-match next month's upload to
    this month's recipe (PRD section 3).

    Built from the primary sheet's column names, their inferred types and the
    header position -- deliberately not from the filename, the row count or the
    values, all of which change every month by design. Two consecutive monthly
    exports from the same system produce the same signature; a different report
    does not.
    """
    if not tables:
        return "empty"

    primary = tables[0].interpretation
    parts = [
        f"header@{primary.header_row}",
        f"cols={primary.last_column - primary.first_column + 1}",
    ]
    for column in primary.columns:
        parts.append(f"{column.name}:{column.inferred_type}")

    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()
    return digest[:32]


def parse_workbook(data: bytes, filename: str) -> ParseResult:
    """Entry point: bytes in, interpreted tables out."""
    grids = read_grids(data, filename)
    if not grids:
        raise ValueError("The workbook has no visible sheets")

    tables: list[ParsedTable] = []
    warnings: list[str] = []

    for sheet_name, grid in grids:
        if not grid:
            continue
        interpretation, data_indices = detect_table(grid, sheet_name)
        if interpretation.header_row is None or not data_indices:
            warnings.append(f"sheet {sheet_name!r}: no table found, skipped")
            continue
        tables.append(build_table(grid, interpretation, data_indices))

    if not tables:
        raise ValueError(
            "No table could be located in this file. Check that it contains a header row "
            "with data beneath it."
        )

    # The primary table is the one with the most data, not the first sheet. The
    # fixture's "Notes" sheet is a real table and would win on position.
    primary_index = max(
        range(len(tables)),
        key=lambda i: (tables[i].row_count, tables[i].interpretation.confidence),
    )
    if primary_index != 0:
        warnings.append(
            f"sheet {tables[primary_index].interpretation.sheet_name!r} treated as the main table"
        )

    ordered = [tables[primary_index]] + [t for i, t in enumerate(tables) if i != primary_index]

    for table in ordered:
        low = [c for c in table.interpretation.columns if c.type_confidence < 0.8
               and c.inferred_type in {"date", "number"}]
        for column in low:
            warnings.append(
                f"{table.interpretation.sheet_name}.{column.name}: "
                f"{column.inferred_type} inferred at {column.type_confidence:.0%} confidence"
            )
        for column in table.interpretation.columns:
            if column.ambiguous_dates:
                warnings.append(
                    f"{table.interpretation.sheet_name}.{column.name}: "
                    f"{column.ambiguous_dates} date(s) could be read either day-first or month-first; "
                    f"read as {column.date_order}"
                )

    return ParseResult(
        tables=ordered,
        primary_index=0,
        source_signature=compute_source_signature(ordered),
        warnings=warnings,
    )


__all__ = [
    "ColumnInterpretation",
    "ParseResult",
    "ParsedTable",
    "SheetInterpretation",
    "SkippedRow",
    "compute_source_signature",
    "detect_table",
    "entity_key",
    "parse_workbook",
    "read_grids",
]
