"""
Report generation.

What a report *says* is decided here, once, as a `ReportDocument` of typed
blocks. How it looks is decided four times: `render_markdown` below, and PDF,
Word and Excel in `documents.py`.

Markdown remains the default, and the reasoning that made it the only format
still holds for the copy that goes into the working papers: it survives an
email, a client portal and a `git diff`, and anybody can edit it. What that
reasoning missed is the other reader. A month-end pack sent *to the client* is
a document with the firm's name on it, and handing that reader a .md file asks
them to care about our tooling. The three binary renderings exist for them, and
for the sales case that turns on producing a client-facing report in seconds
that a firm currently builds by hand.

Every figure in a report comes from `analyze`, and every one of them carries
the row count behind it. Section 7's promise is that any displayed number can
be traced, so a report that states a total without saying how many rows it
covers is already outside the design -- in any of the four formats.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from dataclasses import dataclass
from typing import Any


# Column names that a naive .title() mangles. An accountant reading "Vat" in a
# report they are about to send to a client notices immediately.
_ACRONYMS = {"vat": "VAT", "gbp": "GBP", "usd": "USD", "eur": "EUR", "id": "ID", "po": "PO"}


def _money(value: float | None, symbol: str = "£") -> str:
    """
    Parentheses for negatives, which is what a set of accounts uses.

    The source files already write credit notes as `(150.00)`; rendering them
    back as `-£150.00` would be correct and still read as foreign to the person
    checking the report against their own spreadsheet.
    """
    if value is None:
        return "—"
    if value < 0:
        return f"({symbol}{abs(value):,.2f})"
    return f"{symbol}{value:,.2f}"


def _label(name: str) -> str:
    return " ".join(
        _ACRONYMS.get(word.lower(), word.title()) for word in name.replace("_", " ").split()
    )


def _period_label(period: dict[str, Any]) -> str:
    """
    The reporting period, phrased the way a covering email would phrase it.

    One month collapses to "September 2026"; a range spanning months keeps both
    ends. An absent or unparseable range returns "", and every renderer treats
    that as "print nothing" rather than as a gap to fill -- a header reading
    "Period: unknown" is worse than a header that does not mention one.
    """
    earliest, latest = period.get("earliest"), period.get("latest")
    if not isinstance(earliest, str) or not isinstance(latest, str):
        return ""
    try:
        start = dt.date.fromisoformat(earliest[:10])
        end = dt.date.fromisoformat(latest[:10])
    except ValueError:
        return ""
    if (start.year, start.month) == (end.year, end.month):
        return start.strftime("%B %Y")
    if start.year == end.year:
        return f"{start.strftime('%B')} to {end.strftime('%B %Y')}"
    return f"{start.strftime('%B %Y')} to {end.strftime('%B %Y')}"


def _table(headers: list[str], rows: list[list[str]]) -> str:
    if not rows:
        return "_No rows._\n"
    lines = [
        "| " + " | ".join(headers) + " |",
        "|" + "|".join("---" for _ in headers) + "|",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in rows)
    return "\n".join(lines) + "\n"


def _bar_chart(rows: list[tuple[str, float]], width: int = 32) -> str:
    """
    A bar chart drawn in text.

    Deliberately not a PNG. This report is Markdown because an accountant's
    month-end pack goes into an email, a working paper or a client portal, and
    an image would either travel as a second file that gets separated from it or
    as a base64 blob that half of those viewers refuse to render. A bar made of
    block characters survives all of them, including a plain-text paste, and
    needs no plotting dependency on the agent host.

    Scaled against the largest absolute value so a month of credits reads as a
    bar on the same axis rather than as an empty row. The number is always
    printed beside the bar: the bar is for the shape, the figure is the fact.
    """
    if not rows:
        return ""

    largest = max(abs(value) for _label, value in rows) or 1.0
    longest_label = max(len(label) for label, _value in rows)

    lines = ["```"]
    for label, value in rows:
        filled = int(round(abs(value) / largest * width))
        bar = ("█" * filled).ljust(width, "░")
        lines.append(f"{label.ljust(longest_label)}  {bar}  {_money(value)}")
    lines.append("```")
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# The document model
#
# One report, four renderings. The assembly below decides *what* a month-end
# report says -- which figures, in which order, with which caveats -- and says
# it once, in blocks that carry meaning rather than formatting. Markdown,
# PDF, Word and Excel are then four functions that each know how to draw a
# heading, a table and a bar, and nothing at all about which tables a report
# contains.
#
# The alternative -- a build_pdf_report() beside build_markdown_report() --
# was rejected on sight. The reconciliation warning at the top is the clearest
# reason: it is a compliance behaviour, not a decoration, and a second copy of
# the assembly is a second place for it to be forgotten.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Figure:
    """
    One headline number.

    `note` is the qualifier that belongs *under* the value rather than beside
    it -- "8 credit rows" under a total. Markdown has nowhere to put that, so
    it renders as its own row; the three visual formats set it in small type
    inside the card.
    """

    label: str
    value: str
    note: tuple[str, str] | None = None


@dataclass(frozen=True)
class Heading:
    text: str


@dataclass(frozen=True)
class Prose:
    text: str


@dataclass(frozen=True)
class Callout:
    """
    Something the reader must see before the figures, not after them.

    `tone` is one of `danger`, `warning`, `info` -- the same three words the
    dashboard uses, so a report and the screen it came from never disagree
    about how serious something is.
    """

    title: str
    lines: list[str]
    tone: str = "danger"


@dataclass(frozen=True)
class KeyFigures:
    figures: list[Figure]


@dataclass(frozen=True)
class Table:
    """
    `numeric` names the columns that hold money or counts, by index.

    Rendering does not care what a column means, but it does care whether the
    digits line up: a total right-aligned under a total is readable, and the
    same column left-aligned is not. The builder knows which is which, so it
    says so here rather than leaving each renderer to guess from the content.
    """

    headers: list[str]
    rows: list[list[str]]
    numeric: tuple[int, ...] = ()


@dataclass(frozen=True)
class Bars:
    """
    A ranked comparison. `labels` holds the already-formatted value beside each
    bar, parallel to `items`, because the formatting rules for money live in
    one place and it is not the renderer.
    """

    items: list[tuple[str, float]]
    labels: list[str]


@dataclass(frozen=True)
class Footnote:
    text: str


Block = Heading | Prose | Callout | KeyFigures | Table | Bars | Footnote


@dataclass(frozen=True)
class ReportDocument:
    title: str
    workspace_name: str
    version_no: int
    generated_at: dt.datetime
    blocks: list[Block]
    #: What the figures cover, as a person would say it -- "September 2026",
    #: or "January to March 2026". Derived from the data's own date range
    #: rather than from the clock, because a report produced in October about
    #: September is a September report, and dating it by when it was rendered
    #: is how a month-end pack ends up filed under the wrong month.
    period: str = ""


def build_report_document(
    *,
    workspace_name: str,
    dataset_name: str,
    version_no: int,
    kpis: dict[str, Any],
    profile_signals: dict[str, Any],
    proposals_summary: dict[str, Any] | None = None,
    comparison: dict[str, Any] | None = None,
    provenance: dict[str, Any] | None = None,
    narrative: str | None = None,
    generated_at: dt.datetime | None = None,
) -> ReportDocument:
    """
    Everything a month-end report says, in the order it says it.
    """
    generated = generated_at or dt.datetime.now(dt.timezone.utc)
    blocks: list[Block] = []
    period_label = _period_label(kpis.get("period") or {})

    # The caveat goes at the top, not in a footnote. A report whose totals do
    # not reconcile against the source file must say so before anyone reads the
    # totals -- section 5.3's blocking behaviour, carried into the document.
    totals = profile_signals.get("declared_totals", {})
    if totals.get("checked") and totals.get("all_reconcile") is False:
        lines = [
            f"{check['column']}: computed {_money(check['computed'])} against a declared "
            f"{_money(check['declared'])} — a difference of {_money(check['difference'])}."
            for check in totals.get("checks", [])
            if not check["reconciles"]
        ]
        lines.append("Resolve this before the figures are relied on.")
        blocks.append(
            Callout("These figures do not reconcile to the source file.", lines, tone="danger")
        )

    if narrative:
        blocks.append(Heading("Summary"))
        blocks.append(Prose(narrative))

    blocks.append(Heading("Headline figures"))
    figures: list[Figure] = [Figure("Rows", f"{kpis.get('row_count', 0):,}")]

    period = kpis.get("period") or {}
    if period.get("earliest"):
        figures.append(Figure("Period", f"{period['earliest']} to {period['latest']}"))

    for key, value in kpis.items():
        if not isinstance(value, dict) or "total" not in value:
            continue
        note = (
            ("— of which credits", f"{value['negative_rows']} row(s)")
            if value.get("negative_rows")
            else None
        )
        figures.append(Figure(_label(key), _money(value["total"]), note))

    blocks.append(KeyFigures(figures))

    monthly = kpis.get("monthly") or {}
    series = monthly.get("series") or []
    if len(series) > 1:
        blocks.append(Heading(f"{_label(monthly['metric'])} by month"))
        blocks.append(
            Bars(
                [(item["month"], item["total"]) for item in series],
                [_money(item["total"]) for item in series],
            )
        )

    for key, value in kpis.items():
        if not key.startswith("top_by_") or not isinstance(value, list):
            continue
        dimension = key.removeprefix("top_by_").replace("_", " ")
        blocks.append(Heading(f"By {dimension}"))
        blocks.append(
            Table(
                [_label(dimension), "Total", "Rows"],
                [
                    [str(item["label"]), _money(item["total"]), str(item["rows"])]
                    for item in value
                ],
                numeric=(1, 2),
            )
        )
        blocks.append(
            Bars(
                [(str(item["label"])[:28], item["total"]) for item in value[:8]],
                [_money(item["total"]) for item in value[:8]],
            )
        )

    if comparison:
        blocks.append(Heading("Period comparison"))
        change = comparison.get("percent_change")
        change_text = f"{change:+.1f}%" if change is not None else "n/a"
        blocks.append(
            Prose(
                f"{comparison['metric']}: {_money(comparison['total_a'])} → "
                f"{_money(comparison['total_b'])} "
                f"({_money(comparison['difference'])}, {change_text})"
            )
        )
        if comparison.get("drivers"):
            blocks.append(Prose("Largest movements:"))
            blocks.append(
                Table(
                    ["", "Previous", "Current", "Change"],
                    [
                        [
                            str(driver["label"]),
                            _money(driver["period_a"]),
                            _money(driver["period_b"]),
                            _money(driver["difference"]),
                        ]
                        for driver in comparison["drivers"]
                    ],
                    numeric=(1, 2, 3),
                )
            )

    blocks.append(Heading("Data quality"))
    quality_rows: list[list[str]] = []

    duplicates = profile_signals.get("exact_duplicates", {})
    quality_rows.append(["Exact duplicate rows", str(duplicates.get("duplicate_rows", 0))])

    variants = profile_signals.get("entity_variants", {}).get("columns", [])
    quality_rows.append(
        ["Name-variant groups", str(sum(item["group_count"] for item in variants))]
    )

    vat = profile_signals.get("vat_consistency", {})
    if vat.get("checked"):
        quality_rows.append(["VAT rate anomalies", str(vat.get("anomaly_count", 0))])
        quality_rows.append(["VAT rates seen", str(vat.get("rate_distribution", {}))])

    dates = profile_signals.get("date_coverage", {})
    if dates.get("checked") and dates.get("ambiguous_dates"):
        quality_rows.append(
            [
                "Ambiguous dates",
                f"{dates['ambiguous_dates']} read as {str(dates.get('assumed_order', '')).upper()}",
            ]
        )

    if proposals_summary:
        quality_rows.append(
            ["Changes applied automatically", str(proposals_summary.get("auto", 0))]
        )
        quality_rows.append(["Changes needing review", str(proposals_summary.get("review", 0))])
        quality_rows.append(
            ["Value under review", _money(proposals_summary.get("review_materiality_gbp"))]
        )

    blocks.append(Table(["Check", "Result"], quality_rows, numeric=(1,)))

    # Where these figures came from and what was done to them on the way.
    #
    # The question an accountant asks before signing anything is not "what is
    # the total" but "which file is this, and what changed before you counted".
    # Every fact needed to answer that was already in the database -- the source
    # workbook, the version chain, the changes a person approved -- and the
    # report simply never asked for any of it.
    if provenance:
        blocks.append(Heading("Provenance"))
        trail: list[list[str]] = []

        if provenance.get("source_filename"):
            trail.append(["Source workbook", str(provenance["source_filename"])])
        if provenance.get("uploaded_at"):
            trail.append(["Uploaded", str(provenance["uploaded_at"])[:19].replace("T", " ")])
        trail.append(["Dataset version", f"v{version_no}"])
        if provenance.get("parsed_directly"):
            trail.append(["Produced by", "reading the source workbook"])
        elif provenance.get("parent_version_no") is not None:
            trail.append(["Derived from", f"v{provenance['parent_version_no']}"])
        if provenance.get("row_count") is not None:
            trail.append(["Rows in this version", f"{provenance['row_count']:,}"])

        blocks.append(Table(["", ""], trail))

        applied = provenance.get("applied_changes") or []
        if applied:
            blocks.append(
                Prose(
                    f"{len(applied)} change(s) were approved by a person and applied to produce "
                    f"this version:"
                )
            )
            blocks.append(
                Table(
                    ["Change", "Rows", "Decided"],
                    [
                        [
                            str(change.get("title", "")),
                            str(change.get("affected_rows", 0)),
                            str(change.get("decided_at", ""))[:10],
                        ]
                        for change in applied
                    ],
                    numeric=(1,),
                )
            )
        elif provenance.get("parent_version_no") is not None:
            blocks.append(Prose("No changes were applied to produce this version."))

    # The claim is deliberately narrower than it was. It used to read "every
    # figure above is computed from the stored dataset" -- which is true of the
    # tables and not of the summary, because that paragraph is written by a
    # model from the profile. The numbers in it have been right in testing;
    # nothing in the code makes them right, and a blanket guarantee over
    # unvalidated prose is exactly the sentence that matters after something
    # goes wrong.
    claim = "Every figure in the tables above is computed from the stored dataset, not estimated."
    if narrative:
        claim += (
            " The summary is written by a language model from those same figures and is not "
            "independently checked."
        )

    blocks.append(
        Footnote(
            f"Produced by the Hermes agent from dataset version {version_no}. {claim} "
            "A copilot, not an autonomous accountant — review before use."
        )
    )

    return ReportDocument(
        title=dataset_name,
        workspace_name=workspace_name,
        version_no=version_no,
        generated_at=generated,
        blocks=blocks,
        period=period_label,
    )


def render_markdown(document: ReportDocument, brand: Any = None) -> str:
    """
    The plain-text rendering, and still the default one.

    An accountant's month-end pack goes into an email, a working paper or a
    client portal, and Markdown survives all three. The three binary formats in
    `documents.py` exist for the copy that goes *to the client*, where the
    layout is part of the product; this one is for the copy that goes into the
    file, where being editable and diffable matters more than being beautiful.

    `brand` is duck-typed rather than imported. `documents.Brand` lives in the
    module that imports this one, and a report that renders as text must not
    depend on the module that needs reportlab -- typing the parameter properly
    would buy a cycle and an import of three rendering libraries to produce a
    string.

    Markdown cannot show a private logo (section 18), and this deliberately does
    not try. A signed storage URL expires and is a credential; what appears here
    is the business name, and an image only when an administrator has supplied a
    public URL for one.
    """
    parts: list[str] = []
    business_name = getattr(brand, "name", None) if brand is not None else None

    if business_name:
        logo_url = getattr(brand, "logo_url", None)
        if isinstance(logo_url, str) and logo_url.startswith("https://"):
            parts.append(f"![{business_name}]({logo_url})\n")
        parts.append(f"# {business_name}\n")
        parts.append(f"## {document.title}\n")
    else:
        parts.append(f"# {document.title}\n")

    if document.period:
        parts.append(f"{document.period}\n")

    parts.append(
        f"**{document.workspace_name}** · dataset version {document.version_no} · "
        f"generated {document.generated_at.strftime('%d %B %Y %H:%M UTC')}\n"
    )

    for block in document.blocks:
        if isinstance(block, Heading):
            parts.append(f"\n## {block.text}\n")

        elif isinstance(block, Prose):
            parts.append(f"{block.text}\n")

        elif isinstance(block, Callout):
            # One blockquote, joined here rather than appended line by line: a
            # trailing `>` after the last line renders as an empty quoted
            # paragraph in every viewer that takes Markdown seriously.
            quoted = [f"> **{block.title}**", ">"]
            for line in block.lines:
                quoted.extend([f"> {line}", ">"])
            parts.append("\n" + "\n".join(quoted[:-1]) + "\n")

        elif isinstance(block, KeyFigures):
            rows: list[list[str]] = []
            for figure in block.figures:
                rows.append([figure.label, figure.value])
                if figure.note:
                    rows.append([figure.note[0], figure.note[1]])
            parts.append(_table(["Measure", "Value"], rows))

        elif isinstance(block, Table):
            parts.append(_table(block.headers, block.rows))

        elif isinstance(block, Bars):
            parts.append(_bar_chart(block.items))

        elif isinstance(block, Footnote):
            parts.append(f"\n---\n\n_{block.text}_\n")

    footer = getattr(brand, "footer", None) if brand is not None else None
    if isinstance(footer, str) and footer.strip():
        parts.append(f"\n_{footer.strip()}_\n")

    return "\n".join(parts)


def build_markdown_report(
    *,
    workspace_name: str,
    dataset_name: str,
    version_no: int,
    kpis: dict[str, Any],
    profile_signals: dict[str, Any],
    proposals_summary: dict[str, Any] | None = None,
    comparison: dict[str, Any] | None = None,
    provenance: dict[str, Any] | None = None,
    narrative: str | None = None,
    generated_at: dt.datetime | None = None,
) -> str:
    """Kept as the one-call entry point every caller already uses."""
    return render_markdown(
        build_report_document(
            workspace_name=workspace_name,
            dataset_name=dataset_name,
            version_no=version_no,
            kpis=kpis,
            profile_signals=profile_signals,
            proposals_summary=proposals_summary,
            comparison=comparison,
            provenance=provenance,
            narrative=narrative,
            generated_at=generated_at,
        )
    )


def rows_to_csv(rows: list[dict[str, Any]]) -> bytes:
    """Export helper for the `exports` bucket."""
    if not rows:
        return b""

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()), extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    # BOM so Excel opens it as UTF-8 rather than mangling the pound sign, which
    # is the first thing anyone will notice in a UK accounting export.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


# Types openpyxl writes natively. Everything else -- a dict from a nested
# column, a Decimal, a UUID -- becomes its string form, because a cell that
# raises on write loses the whole export over one awkward value.
_XLSX_NATIVE = (str, int, float, bool, dt.datetime, dt.date, dt.time)


def _xlsx_cell(value: Any) -> Any:
    if value is None or isinstance(value, _XLSX_NATIVE):
        return value
    return str(value)


def rows_to_xlsx(rows: list[dict[str, Any]], sheet_name: str = "Data") -> bytes:
    """
    Export helper for the `exports` bucket, for people who open the file rather
    than parse it.

    csv survives everything and xlsx is what actually gets opened, so both
    exist. The difference that matters is types: a csv hands Excel a pile of
    text and lets it guess, which is how an account code of 0041 becomes 41 and
    how 03/04 becomes a date in March. Writing real cell types means the values
    arrive as the parser understood them.

    openpyxl is imported here rather than at module scope because the markdown
    path -- which every report job runs -- has no use for it.
    """
    # Imported lazily: see the docstring.
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    # Excel refuses >31 chars and the characters below; a dataset named after a
    # client file will hit both.
    sheet.title = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Data"

    if rows:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append([_xlsx_cell(row.get(header)) for header in headers])

        # The header stays put while someone scrolls a few thousand rows. Cheap,
        # and its absence is the first thing anyone notices.
        sheet.freeze_panes = "A2"

        for index, header in enumerate(headers, start=1):
            width = max(len(str(header)), *(len(str(row.get(header, ""))) for row in rows))
            sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = [
    "Bars",
    "Block",
    "Callout",
    "Figure",
    "Footnote",
    "Heading",
    "KeyFigures",
    "Prose",
    "ReportDocument",
    "Table",
    "build_markdown_report",
    "build_report_document",
    "render_markdown",
    "rows_to_csv",
    "rows_to_xlsx",
]
